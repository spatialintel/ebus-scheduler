"""
output_formatter.py — Produces the final Bus Schedule Excel.

Output columns:
    Route | UP/DN | Start Location | End Location | Bus Number |
    Start Time | End Time | Break (min) | Distance (km) | Battery (kWh) |
    SOC (%) | Terminal Charging (min) | Depot Charging (min) | Shift | Type

Usage:
    from src.output_formatter import write_schedule
    write_schedule(config, buses, "outputs/R1_schedule.xlsx")
"""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models import BusState, RouteConfig


# ── Styles ──
HDR_FILL = PatternFill('solid', fgColor='2F4F4F')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY_FONT = Font(name='Arial', size=9)
DEAD_FILL = PatternFill('solid', fgColor='F5F5F5')
CHARGE_FILL = PatternFill('solid', fgColor='E8F5E9')
LOW_SOC_FONT = Font(name='Arial', size=9, color='CC0000', bold=True)
THIN = Border(
    left=Side('thin', color='DDDDDD'), right=Side('thin', color='DDDDDD'),
    top=Side('thin', color='DDDDDD'), bottom=Side('thin', color='DDDDDD'),
)

COLUMNS = [
    ("Route",                8),
    ("UP/DN",               10),
    ("Start Location",      28),
    ("End Location",        28),
    ("Bus",                  6),
    ("Start Time",          12),
    ("End Time",            12),
    ("Break (min)",         12),
    ("Distance (km)",       13),
    ("Battery (kWh)",       13),
    ("SOC (%)",             10),
    ("Terminal Chg (min)",  17),
    ("Depot Chg (min)",     15),
    ("Shift",                6),
    ("Type",                10),
]


def _build_rows(config: RouteConfig, buses: list[BusState]) -> list[dict]:
    """
    Flatten all bus trips into a sorted row list with computed fields.
    """
    rows = []

    for bus in buses:
        running_soc_kwh = config.initial_soc_percent / 100 * config.battery_kwh

        for i, trip in enumerate(bus.trips):
            # Energy consumed on this trip
            energy_used = trip.distance_km * config.consumption_rate
            running_soc_kwh -= energy_used
            soc_pct = (running_soc_kwh / config.battery_kwh) * 100

            # If this is a charging trip, add energy back
            depot_chg_min = 0
            terminal_chg_min = 0
            if trip.trip_type == "Charging":
                depot_chg_min = trip.travel_time_min
                kwh_added = config.depot_flow_rate_kw * (trip.travel_time_min / 60)
                running_soc_kwh = min(config.battery_kwh, running_soc_kwh + kwh_added)
                soc_pct = (running_soc_kwh / config.battery_kwh) * 100

            # Break: gap to next trip
            break_min = None
            if i + 1 < len(bus.trips):
                next_trip = bus.trips[i + 1]
                if trip.actual_arrival and next_trip.actual_departure:
                    gap = (next_trip.actual_departure - trip.actual_arrival).total_seconds() / 60
                    break_min = max(0, int(gap))

            # Direction label
            if trip.trip_type == "Dead" or trip.trip_type == "Charging":
                direction = "DEPOT"
            elif trip.direction == "UP":
                direction = f"{config.route_code}UP"
            else:
                direction = f"{config.route_code}DN"

            # Shift
            shift = trip.shift or (1 if trip.actual_departure and
                    trip.actual_departure.hour < config.shift_split.hour else 2)

            rows.append({
                "route": config.route_code,
                "direction": direction,
                "start_loc": trip.start_location,
                "end_loc": trip.end_location,
                "bus": trip.assigned_bus or bus.bus_id,
                "start_time": trip.actual_departure,
                "end_time": trip.actual_arrival,
                "break_min": break_min,
                "distance_km": trip.distance_km,
                "battery_kwh": round(running_soc_kwh, 1),
                "soc_pct": round(soc_pct, 1),
                "terminal_chg": terminal_chg_min,
                "depot_chg": depot_chg_min,
                "shift": shift,
                "trip_type": trip.trip_type,
                "_sort_key": (trip.actual_departure or trip.earliest_departure,
                              bus.bus_id),
            })

    # Sort by time then bus
    rows.sort(key=lambda r: r["_sort_key"])
    return rows


def write_schedule(
    config: RouteConfig,
    buses: list[BusState],
    output_path: str | Path,
) -> Path:
    """
    Write the final Bus Schedule to an Excel file.
    Returns the output path.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Bus Schedule ──
    ws = wb.active
    ws.title = "Bus Schedule"
    ws.sheet_view.showGridLines = False

    # Headers
    for ci, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    rows = _build_rows(config, buses)

    for ri, row in enumerate(rows, start=2):
        values = [
            row["route"],
            row["direction"],
            row["start_loc"],
            row["end_loc"],
            row["bus"],
            row["start_time"].strftime("%H:%M") if row["start_time"] else "",
            row["end_time"].strftime("%H:%M") if row["end_time"] else "",
            row["break_min"] if row["break_min"] is not None else "",
            round(row["distance_km"], 1),
            row["battery_kwh"],
            row["soc_pct"],
            row["terminal_chg"] if row["terminal_chg"] > 0 else "",
            row["depot_chg"] if row["depot_chg"] > 0 else "",
            row["shift"],
            row["trip_type"],
        ]

        is_dead = row["trip_type"] == "Dead"
        is_charge = row["trip_type"] == "Charging"
        is_low_soc = row["soc_pct"] < config.min_soc_percent + 5

        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if is_dead:
                cell.fill = DEAD_FILL
            elif is_charge:
                cell.fill = CHARGE_FILL

            # Highlight low SOC
            if ci == 11 and is_low_soc:
                cell.font = LOW_SOC_FONT

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Fleet Summary ──
    ws2 = wb.create_sheet("Fleet Summary")
    ws2.sheet_view.showGridLines = False

    summary_headers = ["Bus", "Total Trips", "Revenue Trips", "Dead Runs",
                       "Charging Stops", "Total KM", "Revenue KM", "Dead KM",
                       "Final SOC (%)", "Last Location"]
    for ci, h in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN
        ws2.column_dimensions[get_column_letter(ci)].width = 14

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['J'].width = 28

    for ri, bus in enumerate(buses, start=2):
        rev_trips = [t for t in bus.trips if t.trip_type == "Revenue"]
        dead_trips = [t for t in bus.trips if t.trip_type == "Dead"]
        chg_trips = [t for t in bus.trips if t.trip_type == "Charging"]
        rev_km = sum(t.distance_km for t in rev_trips)
        dead_km = sum(t.distance_km for t in dead_trips)

        vals = [
            bus.bus_id,
            len(bus.trips),
            len(rev_trips),
            len(dead_trips),
            len(chg_trips),
            round(bus.total_km, 1),
            round(rev_km, 1),
            round(dead_km, 1),
            round(bus.soc_percent, 1),
            bus.current_location,
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws2.freeze_panes = "A2"

    wb.save(path)
    return path
