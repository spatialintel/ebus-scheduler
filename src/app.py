"""
app.py — eBus Scheduler Dashboard
"""
from __future__ import annotations
__version__ = "2026-04-09-p2"
import sys, tempfile, math
from pathlib import Path
from datetime import datetime, timedelta, time as dtime

import streamlit as st
import pandas as pd

try:
    import plotly.graph_objects as go
    _PLOTLY_OK = True
except ModuleNotFoundError:
    _PLOTLY_OK = False

sys.path.insert(0, str(Path(__file__).parent))

from src.config_loader import load_config, ConfigError
from src.distance_engine import enrich_distances
from src.trip_generator import generate_trips, check_headway_feasibility
from src.bus_scheduler import schedule_buses, check_compliance
from src.output_formatter import write_schedule
from src.metrics import compute_metrics

# Citywide imports — safe fallback if files not yet deployed
try:
    from src.city_config_loader import load_city_config_from_files
    from src.city_scheduler import schedule_city, _natural_headway, _flat_headway_df
    from src.city_models import CityConfig, CitySchedule, RouteResult
    from src.fleet_analyzer import (
        compute_pvr_all, compute_pvr_slices_all,
        compute_fleet_balance,
    )
    _CITY_OK = True
except Exception:
    _CITY_OK = False

st.set_page_config(page_title="eBus Scheduler", page_icon="🚌", layout="wide",
                   initial_sidebar_state="expanded")

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
  html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
  .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

  /* KPI cards */
  .kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); gap: 12px; margin-bottom: 1.2rem; }
  .kpi { background: #fff; border: 1px solid #e8eaed; border-radius: 10px; padding: 14px 16px; }
  .kpi-val { font-size: 1.65rem; font-weight: 700; color: #1a1a2e; line-height: 1.1; font-family: 'DM Mono', monospace; }
  .kpi-label { font-size: 0.72rem; font-weight: 500; color: #888; text-transform: uppercase; letter-spacing: .05em; margin-top: 4px; }
  .kpi-sub { font-size: 0.68rem; color: #aaa; margin-top: 2px; }
  .kpi-ok .kpi-val { color: #16a34a; }
  .kpi-warn .kpi-val { color: #d97706; }
  .kpi-bad .kpi-val { color: #dc2626; }

  /* Compliance scorecard */
  .rule-row { display: flex; align-items: center; gap: 12px; padding: 10px 14px; border-radius: 8px; margin-bottom: 6px; background: #fafafa; border: 1px solid #f0f0f0; }
  .rule-row.fail { background: #fff5f5; border-color: #fecaca; }
  .rule-row.warn { background: #fffbeb; border-color: #fde68a; }
  .rule-row.pass { background: #f0fdf4; border-color: #bbf7d0; }
  .rule-badge { font-size: 1.1rem; flex-shrink: 0; }
  .rule-name { font-weight: 600; font-size: 0.88rem; color: #222; flex: 1; }
  .rule-detail { font-size: 0.78rem; color: #666; }
  .rule-tag { font-size: 0.7rem; font-weight: 600; padding: 2px 8px; border-radius: 20px; }
  .tag-p { background: #fee2e2; color: #dc2626; }
  .tag-o { background: #dbeafe; color: #2563eb; }

  /* Bus timeline */
  .bus-header { font-size: 1rem; font-weight: 700; color: #1a1a2e; margin: 1rem 0 0.3rem 0; display: flex; align-items: center; gap: 10px; }
  .bus-pill { font-size: 0.72rem; font-weight: 600; padding: 2px 10px; border-radius: 20px; background: #e0e7ff; color: #4338ca; }

  /* Section headers */
  .section-title { font-size: 1.1rem; font-weight: 700; color: #1a1a2e; margin: 1.2rem 0 0.4rem 0; padding-bottom: 6px; border-bottom: 2px solid #e8eaed; }

  div[data-testid="stDownloadButton"] button { background: #4f46e5; color: white; border: none; border-radius: 8px; font-weight: 600; padding: 0.5rem 1.5rem; }
  div[data-testid="stDownloadButton"] button:hover { background: #4338ca; }
  section[data-testid="stSidebar"] .stButton button { background: #4f46e5; color: white !important; border: none; border-radius: 8px; font-weight: 600; width: 100%; }
  section[data-testid="stSidebar"] .stButton button:hover { background: #4338ca; }
</style>
""", unsafe_allow_html=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def kpi(label, value, status="", sub=""):
    cls = f"kpi {('kpi-ok' if status=='ok' else 'kpi-warn' if status=='warn' else 'kpi-bad' if status=='bad' else '')}"
    sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ""
    return f'<div class="{cls}"><div class="kpi-val">{value}</div><div class="kpi-label">{label}</div>{sub_html}</div>'

def _true_driver_break(bus, i):
    """
    Return the idle break minutes after trip i, or None if not a true driver break.

    A true driver break is the gap between two consecutive Revenue trips with
    NO Charging, Dead, or Shuttle trips between them. Gaps that include a
    charging detour (Revenue → Shuttle → Dead → Charging → Dead → Shuttle →
    Revenue) are excluded — those minutes are spent at the depot, not idle.
    """
    if bus.trips[i].trip_type != "Revenue":
        return None
    # Find the next revenue trip
    nxt_rev = next((t for t in bus.trips[i+1:] if t.trip_type == "Revenue"), None)
    if nxt_rev is None:
        return None
    idx_curr = i
    idx_next = bus.trips.index(nxt_rev)
    # Check for any detour trips between the two revenue trips
    between = bus.trips[idx_curr+1:idx_next]
    if any(t.trip_type in ("Charging", "Dead", "Shuttle") for t in between):
        return None
    if bus.trips[i].actual_arrival and nxt_rev.actual_departure:
        return max(0, int((nxt_rev.actual_departure - bus.trips[i].actual_arrival).total_seconds() / 60))
    return None


def build_schedule_df(config, buses):
    rows = []
    for bus in buses:
        soc = config.initial_soc_percent
        for i, trip in enumerate(bus.trips):
            soc -= (trip.distance_km * config.consumption_rate / config.battery_kwh) * 100
            if trip.trip_type == "Charging":
                soc = min(100.0, soc + config.depot_flow_rate_kw * (trip.travel_time_min/60) / config.battery_kwh * 100)
            brk = _true_driver_break(bus, i)
            if trip.trip_type in ("Dead", "Charging", "Shuttle"): direction = "DEPOT"
            elif trip.direction == "UP": direction = f"{config.route_code}UP"
            else: direction = f"{config.route_code}DN"
            rows.append({
                "Bus": trip.assigned_bus or bus.bus_id, "Direction": direction,
                "Type": trip.trip_type, "From": trip.start_location, "To": trip.end_location,
                "Departure": trip.actual_departure.strftime("%H:%M") if trip.actual_departure else "",
                "Arrival": trip.actual_arrival.strftime("%H:%M") if trip.actual_arrival else "",
                "Break (min)": brk, "Distance (km)": round(trip.distance_km, 1),
                "SOC (%)": round(soc, 1), "Shift": trip.shift,
            })
    return pd.DataFrame(rows)

def build_fleet_df(config, buses):
    rows = []
    for bus in buses:
        rev  = [t for t in bus.trips if t.trip_type == "Revenue"]
        shut = [t for t in bus.trips if t.trip_type == "Shuttle"]
        dead = [t for t in bus.trips if t.trip_type == "Dead"]
        chg  = [t for t in bus.trips if t.trip_type == "Charging"]
        first = next((t.actual_departure for t in bus.trips if t.actual_departure), None)
        last  = next((t.actual_arrival for t in reversed(bus.trips) if t.actual_arrival), None)
        rows.append({
            "Bus": bus.bus_id,
            "Revenue Trips": len(rev),
            "Shuttle Trips": len(shut),
            "Dead Runs": len(dead),
            "Charging": len(chg),
            "Total KM": round(bus.total_km, 1),
            "Revenue KM": round(sum(t.distance_km for t in rev), 1),
            "Dead KM": round(sum(t.distance_km for t in dead), 1),
            "Final SOC (%)": round(bus.soc_percent, 1),
            "Start": first.strftime("%H:%M") if first else "—",
            "End": last.strftime("%H:%M") if last else "—",
        })
    return pd.DataFrame(rows)

def build_route_depiction(config, buses):
    rows = []
    for bus in buses:
        soc = config.initial_soc_percent
        for i, trip in enumerate(bus.trips):
            soc -= (trip.distance_km * config.consumption_rate / config.battery_kwh) * 100
            if trip.trip_type == "Charging":
                soc = min(100, soc + (config.depot_flow_rate_kw * trip.travel_time_min / 60) / config.battery_kwh * 100)
            brk = _true_driver_break(bus, i) or 0
            rows.append({
                "Bus": bus.bus_id,
                "Dep": trip.actual_departure.strftime("%H:%M") if trip.actual_departure else "",
                "Arr": trip.actual_arrival.strftime("%H:%M") if trip.actual_arrival else "",
                "From": trip.start_location, "To": trip.end_location,
                "Type": trip.trip_type,
                "Dist": round(trip.distance_km, 1),
                "SOC": round(soc, 1),
                "Break": brk,
                "_dep_dt": trip.actual_departure,
                "_type": trip.trip_type,
                "_dir": trip.direction,
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Bus", "_dep_dt"]).reset_index(drop=True)
    return df

def build_route_diagram(config, buses, selected_bus=None):
    """
    Two-panel diagram:
    Panel A: Route topology (scaled by distance).
    Panel B: Gantt-style schedule — one wide lane per bus, no overlapping.
    """
    from src.bus_scheduler import _nearest_node_from_depot
    from plotly.subplots import make_subplots

    BUS_COLORS = ["#3B5BDB","#C92A2A","#2B8A3E","#E67700",
                  "#7950F2","#0C8599","#C2255C","#5C940D","#A61E4D"]
    DEAD_CLR = "#F08C00"
    CHG_CLR  = "#F59F00"
    ROW_H    = 0.32   # half-height of revenue bar (in y-units)
    ROW_SEP  = 2.6    # vertical space per bus lane

    nearest_name, _, _ = _nearest_node_from_depot(config)

    def _dist(a, b):
        try: return float(config.get_distance(a, b))
        except: return 0.0
    def _ttime(a, b):
        try: return float(config.get_travel_time(a, b))
        except: return 0.0

    route_nodes = [config.start_point]
    for n in getattr(config, "intermediates", []):
        if n and n.strip(): route_nodes.append(n.strip())
    if config.end_point not in route_nodes:
        route_nodes.append(config.end_point)

    cum = {route_nodes[0]: 0.0}
    for i in range(1, len(route_nodes)):
        f, t = route_nodes[i-1], route_nodes[i]
        cum[t] = cum[f] + _dist(f, t)
    total_km = max(cum.values()) or 1.0
    depot_x  = cum.get(nearest_name, 0.0)
    d2n = _dist(config.depot, nearest_name)
    t2n = _ttime(config.depot, nearest_name)

    op_s = config.operating_start.hour + config.operating_start.minute/60
    op_e = config.operating_end.hour   + config.operating_end.minute/60

    def _socs(bus):
        soc, out = config.initial_soc_percent, []
        for t in bus.trips:
            soc -= t.distance_km * config.consumption_rate / config.battery_kwh * 100
            if t.trip_type == "Charging":
                soc = min(100.0, soc + config.depot_flow_rate_kw
                          * t.travel_time_min / 60 / config.battery_kwh * 100)
            out.append(round(max(0, soc), 1))
        return out

    def _soc_clr(s):
        return "#2B8A3E" if s >= 50 else "#E67700" if s >= 30 else "#C92A2A"

    total_dead = sum(t.distance_km for b in buses for t in b.trips if t.trip_type=="Dead")
    total_rev  = sum(1 for b in buses for t in b.trips if t.trip_type=="Revenue")
    total_shut = sum(1 for b in buses for t in b.trips if t.trip_type=="Shuttle")
    total_chg  = sum(1 for b in buses for t in b.trips if t.trip_type=="Charging")
    avg_soc    = sum(b.soc_percent for b in buses) / max(len(buses), 1)
    n_buses    = len(buses)

    def bus_y(idx):
        return (n_buses - idx) * ROW_SEP

    # ── Figure ────────────────────────────────────────────────────────────────
    fig = make_subplots(
        rows=2, cols=1,
        row_heights=[0.18, 0.82],
        vertical_spacing=0.04,
        subplot_titles=[
            f"<b>Route {config.route_code}</b>  "
            f"<span style=\'color:#6B7280;font-weight:normal\'>{config.route_name} — topology</span>",
            f"<b>Bus Schedule</b>  "
            "<span style=\'font-size:10px;color:#9CA3AF\'>hover bars for trip details</span>",
        ],
    )

    # ════ Panel A — Topology ═════════════════════════════════════════════════
    xpad = total_km * 0.06
    fig.add_trace(go.Scatter(
        x=[cum[n] for n in route_nodes], y=[1]*len(route_nodes),
        mode="lines", line=dict(color="#4C6EF5", width=8),
        hoverinfo="skip", showlegend=False,
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=[depot_x, depot_x], y=[-0.35, 0.78],
        mode="lines", line=dict(color=DEAD_CLR, width=2, dash="dot"),
        hoverinfo="skip", showlegend=False,
    ), row=1, col=1)
    for i in range(len(route_nodes)-1):
        f, t = route_nodes[i], route_nodes[i+1]
        d, tt = _dist(f,t), _ttime(f,t)
        if not d: continue
        mid = (cum[f]+cum[t])/2
        fig.add_annotation(x=mid, y=1.38,
            text=f"<b>{d:.1f} km</b>  ·  {tt:.0f} min",
            showarrow=False, row=1, col=1,
            font=dict(size=11, color="#374151"),
            bgcolor="rgba(255,255,255,0.92)", borderpad=3)
    if d2n:
        fig.add_annotation(x=depot_x + total_km*0.03, y=0.22,
            text=f"<b>{d2n:.1f} km</b>  ·  {t2n:.0f} min",
            showarrow=False, row=1, col=1,
            font=dict(size=10, color=DEAD_CLR),
            bgcolor="rgba(255,255,255,0.9)", borderpad=3)
    for node in route_nodes:
        is_term = node in (config.start_point, config.end_point)
        is_near = node == nearest_name
        clr = "#4C6EF5" if is_term else "#2B8A3E" if is_near else "#0C8599"
        sym = "circle" if is_term else "diamond" if is_near else "square"
        lbl = node.replace("BUS STAND","BS").replace("  "," ").strip()
        fig.add_trace(go.Scatter(
            x=[cum[node]], y=[1], mode="markers+text",
            marker=dict(symbol=sym, size=24 if is_term else 18, color=clr,
                        line=dict(color="white", width=2)),
            text=[f"<b>{lbl}</b>"], textposition="bottom center",
            textfont=dict(size=11, color=clr),
            hovertemplate=f"<b>{node}</b>" +
                ("<br><i>Nearest depot node (P2)</i>" if is_near else "") + "<extra></extra>",
            showlegend=False,
        ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=[depot_x], y=[-0.35], mode="markers+text",
        marker=dict(symbol="square", size=24, color=DEAD_CLR,
                    line=dict(color="white", width=2)),
        text=["<b>DEPOT</b>"], textposition="bottom center",
        textfont=dict(size=11, color=DEAD_CLR),
        hovertemplate=f"<b>{config.depot}</b><br>Dead km total: {total_dead:.1f}<extra></extra>",
        showlegend=False,
    ), row=1, col=1)
    fig.add_annotation(
        x=total_km/2, y=1.85,
        text=(f"<b>{config.fleet_size} buses</b>  ·  "
              f"<b>{total_rev}</b> revenue  +  <b>{total_shut}</b> shuttle  ·  "
              f"Dead {total_dead:.1f} km  ·  {total_chg} charge stops  ·  "
              f"Avg final SOC {avg_soc:.0f}%"),
        showarrow=False, row=1, col=1,
        font=dict(size=12, color="#1F2937"),
        bgcolor="rgba(238,242,255,0.97)",
        borderpad=8, bordercolor="#4C6EF5", borderwidth=1,
    )
    fig.update_xaxes(range=[-xpad, total_km+xpad], showgrid=False, zeroline=False,
                     tickformat=".1f", ticksuffix=" km",
                     title_text="Cumulative distance (km)",
                     color="#9CA3AF", row=1, col=1)
    fig.update_yaxes(range=[-1.0, 2.15], showgrid=False, zeroline=False,
                     showticklabels=False, showline=False, row=1, col=1)

    # ════ Panel B — Gantt ═════════════════════════════════════════════════════

    # Hour gridlines
    for h in range(int(op_s), int(op_e)+1):
        fig.add_shape(type="line", x0=h, x1=h,
                      y0=ROW_SEP*0.3, y1=(n_buses+0.4)*ROW_SEP,
                      line=dict(color="rgba(0,0,0,0.05)", width=1),
                      xref="x2", yref="y2")

    # Peak shading
    for x0, x1, lbl, clr in [
        (8, 11,  "Peak AM",  "rgba(99,102,241,0.06)"),
        (11, 15, "Off-peak", "rgba(16,185,129,0.04)"),
        (16, 20, "Peak PM",  "rgba(99,102,241,0.06)"),
    ]:
        fig.add_vrect(x0=x0, x1=x1, fillcolor=clr, line_width=0, row=2, col=1)
        fig.add_annotation(
            x=(x0+x1)/2, y=(n_buses+0.45)*ROW_SEP,
            xref="x2", yref="y2",
            text=f"<span style=\'font-size:10px;color:#9CA3AF\'><b>{lbl}</b></span>",
            showarrow=False, xanchor="center",
        )

    legend_shown = set()

    for bidx, bus in enumerate(buses):
        bclr   = BUS_COLORS[bidx % len(BUS_COLORS)]
        yc     = bus_y(bidx)
        is_sel = selected_bus is None or bus.bus_id == selected_bus
        alpha  = 1.0 if is_sel else 0.15
        socs   = _socs(bus)

        # Lane background stripe — alternating subtle color
        stripe_clr = "rgba(248,249,250,0.9)" if bidx % 2 == 0 else "rgba(255,255,255,0)"
        fig.add_shape(type="rect",
                      x0=op_s-0.3, x1=op_e+0.3,
                      y0=yc - ROW_SEP*0.45, y1=yc + ROW_SEP*0.45,
                      fillcolor=stripe_clr, line_width=0,
                      xref="x2", yref="y2")

        # Bus label left
        fig.add_annotation(
            x=op_s - 0.22, y=yc, xref="x2", yref="y2",
            text=f"<b style=\'color:{bclr}\' >{bus.bus_id}</b>",
            showarrow=False, xanchor="right",
            font=dict(size=14, color=bclr),
        )

        # Collect revenue trip pairs to annotate breaks BELOW lane (no overlap)
        rev_trips = []

        for tidx, (trip, soc_after) in enumerate(zip(bus.trips, socs)):
            if trip.actual_departure is None: continue
            dep_h = trip.actual_departure.hour + trip.actual_departure.minute/60
            arr_h = (trip.actual_arrival.hour  + trip.actual_arrival.minute/60
                     if trip.actual_arrival else dep_h + trip.travel_time_min/60)
            sc    = _soc_clr(soc_after)
            deps  = trip.actual_departure.strftime("%H:%M")
            arrs  = trip.actual_arrival.strftime("%H:%M") if trip.actual_arrival else "?"
            hover = (f"<b>{bus.bus_id}</b> — {trip.trip_type}<br>"
                     f"{deps} → {arrs}<br>"
                     f"{trip.start_location} → {trip.end_location}<br>"
                     f"SOC after: <b style=\'color:{sc}\'>{soc_after}%</b>"
                     f"  ·  {trip.distance_km:.1f} km<extra></extra>")

            show_leg = bus.bus_id not in legend_shown and is_sel

            if trip.trip_type == "Revenue":
                rev_trips.append((dep_h, arr_h, trip))
                # Main bar
                fig.add_shape(type="rect",
                    x0=dep_h, x1=arr_h,
                    y0=yc - ROW_H, y1=yc + ROW_H,
                    fillcolor=bclr, opacity=0.88*alpha, line_width=0,
                    xref="x2", yref="y2")
                # Direction arrow label centred in bar
                arrow = "▲  UP" if trip.direction == "UP" else "▼  DN"
                if arr_h - dep_h > 0.55:
                    fig.add_annotation(
                        x=(dep_h+arr_h)/2, y=yc, xref="x2", yref="y2",
                        text=f"<span style=\'color:white;font-size:10px\'><b>{arrow}</b></span>",
                        showarrow=False, xanchor="center", yanchor="middle",
                    )
                # Departure circle
                fig.add_shape(type="circle",
                    x0=dep_h-0.04, x1=dep_h+0.04,
                    y0=yc-0.12, y1=yc+0.12,
                    fillcolor="white", line=dict(color=bclr, width=2),
                    xref="x2", yref="y2")
                # Invisible hover scatter
                fig.add_trace(go.Scatter(
                    x=[(dep_h+arr_h)/2], y=[yc],
                    mode="markers",
                    marker=dict(size=1, color=bclr, opacity=0.01),
                    name=bus.bus_id, legendgroup=bus.bus_id,
                    showlegend=show_leg,
                    hovertemplate=hover,
                ), row=2, col=1)
                if show_leg: legend_shown.add(bus.bus_id)

            elif trip.trip_type == "Charging":
                # Amber bar, slightly narrower
                fig.add_shape(type="rect",
                    x0=dep_h, x1=arr_h,
                    y0=yc - ROW_H*0.65, y1=yc + ROW_H*0.65,
                    fillcolor=CHG_CLR, opacity=0.95*alpha,
                    line=dict(color="#92400E", width=1.5),
                    xref="x2", yref="y2")
                dur = int(trip.travel_time_min)
                if arr_h - dep_h > 0.3:
                    fig.add_annotation(
                        x=(dep_h+arr_h)/2, y=yc, xref="x2", yref="y2",
                        text=f"<b style=\'color:#1F2937\'>⚡ {dur}m</b>",
                        showarrow=False, xanchor="center", yanchor="middle",
                        font=dict(size=11),
                    )
                fig.add_trace(go.Scatter(
                    x=[(dep_h+arr_h)/2], y=[yc],
                    mode="markers",
                    marker=dict(size=1, color=CHG_CLR, opacity=0.01),
                    showlegend=False,
                    hovertemplate=hover,
                ), row=2, col=1)

            else:  # Dead or Shuttle — thin dashed connector
                w = 2.5 if trip.trip_type == "Shuttle" else 1.5
                dash = "dash" if trip.trip_type == "Shuttle" else "dot"
                fig.add_trace(go.Scatter(
                    x=[dep_h, arr_h], y=[yc, yc],
                    mode="lines",
                    line=dict(color=DEAD_CLR, width=w, dash=dash),
                    opacity=0.6*alpha,
                    showlegend=False,
                    hovertemplate=hover,
                ), row=2, col=1)

        # Break annotations BELOW the lane — no overlap with bars
        for i in range(1, len(rev_trips)):
            _, arr_h_prev, tp = rev_trips[i-1]
            dep_h_next, _, tc = rev_trips[i]
            # only direct Revenue→Revenue breaks (no Dead/Charge between)
            idx_p = bus.trips.index(tp)
            idx_c = bus.trips.index(tc)
            between = bus.trips[idx_p+1:idx_c]
            if any(t.trip_type in ("Charging","Dead","Shuttle") for t in between):
                continue
            gap = (tc.actual_departure - tp.actual_arrival).total_seconds()/60
            if gap < 2: continue
            mid_h  = (arr_h_prev + dep_h_next) / 2
            clr_b  = "#C92A2A" if gap > 20 else "#495057"
            weight = "bold" if gap > 20 else "normal"
            fig.add_annotation(
                x=mid_h, y=yc - ROW_H - 0.25, xref="x2", yref="y2",
                text=f"<span style=\'color:{clr_b};font-weight:{weight};font-size:10px\'>{gap:.0f}m</span>",
                showarrow=False, xanchor="center", yanchor="top",
            )

        # Final SOC badge — right of lane
        final_soc = socs[-1] if socs else 0
        sclr = _soc_clr(final_soc)
        fig.add_annotation(
            x=op_e + 0.12, y=yc, xref="x2", yref="y2",
            text=f"<b style=\'color:{sclr}\'>{final_soc:.0f}%</b>",
            showarrow=False, xanchor="left", yanchor="middle",
            font=dict(size=12, color=sclr),
        )

    # Legend entries for trip types
    for lbl, clr, sym in [
        ("Revenue trip",   "#4C6EF5", "square"),
        ("Charging ⚡",    CHG_CLR,   "star"),
        ("Dead run",       DEAD_CLR,  "line-ew"),
        ("Shuttle",        "#85C1E9", "line-ew-open"),
    ]:
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(symbol=sym, size=11, color=clr),
            name=lbl, showlegend=True, legendgroup="types",
        ), row=2, col=1)

    # Panel B axes
    tick_vals = list(range(int(op_s), int(op_e)+1))
    tick_text = [f"{h:02d}:00" for h in tick_vals]
    fig.update_xaxes(
        range=[op_s-0.35, op_e+0.55],
        tickvals=tick_vals, ticktext=tick_text,
        showgrid=False, zeroline=False,
        title_text="Time of day",
        color="#6B7280", row=2, col=1,
    )
    fig.update_yaxes(
        range=[ROW_SEP*0.1, (n_buses+0.7)*ROW_SEP],
        showgrid=False, zeroline=False,
        showticklabels=False,
        row=2, col=1,
    )

    panel_b_px = max(120 * n_buses, 520)
    fig.update_layout(
        height=220 + panel_b_px,
        margin=dict(l=70, r=100, t=55, b=30),
        plot_bgcolor="white",
        paper_bgcolor="rgba(0,0,0,0)",
        hovermode="closest",
        legend=dict(
            orientation="v", x=1.01, y=0.5,
            xanchor="left", yanchor="middle",
            font=dict(size=11),
            bgcolor="rgba(255,255,255,0.96)",
            bordercolor="#E5E7EB", borderwidth=1,
            tracegroupgap=3,
        ),
    )
    return fig


def build_headway_chart_data(config, buses):
    rows = []
    for bus in buses:
        for trip in bus.trips:
            if trip.trip_type == "Revenue" and trip.actual_departure:
                rows.append({"Bus": bus.bus_id, "Direction": trip.direction,
                             "Departure": trip.actual_departure})
    return pd.DataFrame(rows)

def _configured_hw_at(t, headway_df):
    """Return configured headway_min for the band containing datetime t."""
    if headway_df is None or headway_df.empty:
        return None
    for _, row in headway_df.iterrows():
        try:
            tf = datetime.strptime(str(row["time_from"]), "%H:%M").time()
            tt = datetime.strptime(str(row["time_to"]),   "%H:%M").time()
            if tf <= t.time() < tt:
                return int(row["headway_min"])
        except Exception:
            continue
    # fallback: last band
    try:
        return int(headway_df.iloc[-1]["headway_min"])
    except Exception:
        return None


def build_headway_fig(deps, headway_df, direction, color, route_label):
    """
    Build a detailed plotly headway bar chart for one direction.

    Features:
    - Bars coloured by severity: green (≤ target), amber (≤ 150% target), red (> 150%)
    - Exact minute value annotated above every bar
    - Step-line overlay showing the configured headway profile
    - Anomaly threshold line at 2× configured headway
    - Gridlines for readability
    """
    if len(deps) < 2:
        return None

    dep_labels, gap_values, cfg_hws = [], [], []
    for i in range(1, len(deps)):
        gap   = round((deps[i] - deps[i-1]).total_seconds() / 60)
        cfg_hw = _configured_hw_at(deps[i], headway_df)
        dep_labels.append(deps[i].strftime("%H:%M"))
        gap_values.append(gap)
        cfg_hws.append(cfg_hw)

    # Colour each bar: green = on-target, amber = moderate deviation, red = anomaly
    bar_colors = []
    for gap, cfg in zip(gap_values, cfg_hws):
        if cfg is None:
            bar_colors.append(color)
        elif gap <= cfg * 1.15:
            bar_colors.append(color)           # on-target
        elif gap <= cfg * 1.5:
            bar_colors.append("#f59e0b")       # amber — notable deviation
        else:
            bar_colors.append("#ef4444")       # red — anomaly / outlier

    fig = go.Figure()

    # Bars
    fig.add_trace(go.Bar(
        x=dep_labels,
        y=gap_values,
        marker_color=bar_colors,
        marker_line_color="rgba(0,0,0,0.15)",
        marker_line_width=0.5,
        text=[f"<b>{v}</b>" for v in gap_values],
        textposition="outside",
        textfont=dict(size=10),
        name="Actual headway",
        hovertemplate="Dep: %{x}<br>Gap: %{y} min<extra></extra>",
        cliponaxis=False,
    ))

    # Configured headway step-line
    if any(c is not None for c in cfg_hws):
        fig.add_trace(go.Scatter(
            x=dep_labels,
            y=cfg_hws,
            mode="lines",
            line=dict(color="#f97316", width=2, dash="dash"),
            name="Configured headway",
            hovertemplate="Configured: %{y} min<extra></extra>",
        ))

    # Anomaly threshold line at 2× median configured headway
    median_cfg = sorted([c for c in cfg_hws if c])[len([c for c in cfg_hws if c])//2] if any(cfg_hws) else None
    if median_cfg:
        fig.add_hline(
            y=median_cfg * 2,
            line_dash="dot", line_color="#ef4444", line_width=1.5,
            annotation_text="Anomaly threshold (2× configured)",
            annotation_position="top right",
            annotation_font=dict(size=10, color="#ef4444"),
        )

    fig.update_layout(
        title=dict(text=f"{direction} headways — {route_label}", font=dict(size=13), x=0),
        xaxis=dict(
            title="Departure time",
            tickangle=-45,
            tickfont=dict(size=9),
            showgrid=True,
            gridcolor="rgba(0,0,0,0.06)",
            gridwidth=1,
        ),
        yaxis=dict(
            title="Gap (min)",
            showgrid=True,
            gridcolor="rgba(0,0,0,0.08)",
            gridwidth=1,
            zeroline=False,
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(size=10)),
        margin=dict(t=60, b=60, l=40, r=20),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        height=300,
        bargap=0.2,
    )
    return fig


def _headway_recommendations(config, headway_df, travel_time_df):
    """
    Compute physics-based recommended headway_min per time band and return
    a DataFrame with columns:
      From, To, Configured, Physics Floor, Recommended, Status, Note

    Logic per band:
      - Look up travel time for the band midpoint (UP direction)
      - cycle_time = 2 × travel_time + 2 × preferred_layover_min
      - natural_gap = cycle_time / fleet_size
      - recommended  = max(configured, ceil(natural_gap)) + 1 buffer
      - Also flag if band headway < recommended (needs change)

    Charging impact row added at the end showing expected max gap during P5.
    """
    import math
    from datetime import datetime as _dt

    if headway_df is None or headway_df.empty:
        return None

    fleet      = max(1, config.fleet_size)
    min_break  = config.preferred_layover_min

    def _tt_for_band(time_from_str):
        """Return best travel time for a band, falling back to segment config."""
        try:
            t = _dt.strptime(str(time_from_str).strip(), "%H:%M")
        except Exception:
            return None
        # Try travel_time_df first
        if travel_time_df is not None and not travel_time_df.empty:
            for _, row in travel_time_df.iterrows():
                try:
                    tf = _dt.strptime(str(row["time_from"]).strip(), "%H:%M")
                    tt = _dt.strptime(str(row["time_to"]).strip(),   "%H:%M")
                    if tf <= t < tt:
                        return float(row.get("up_min", row.get("dn_min", 0)))
                except Exception:
                    continue
        # Fallback: segment config
        try:
            return float(config.get_travel_time(config.start_point, config.end_point))
        except Exception:
            pass
        try:
            dist = config.get_distance(config.start_point, config.end_point)
            spd  = getattr(config, "avg_speed_kmph", 30.0) or 30.0
            return dist / spd * 60
        except Exception:
            return 40.0   # absolute fallback

    # Charging round-trip estimate
    try:
        depot_dead = config.get_travel_time(
            getattr(config, "depot", "DEPOT"),
            config.start_point
        )
    except Exception:
        depot_dead = 30.0
    try:
        cs = config.p5_charging_start
        ce = config.p5_charging_end
        from datetime import datetime as _dt2
        cws = _dt2.strptime(f"{cs.hour:02d}:{cs.minute:02d}", "%H:%M")
        cwe = _dt2.strptime(f"{ce.hour:02d}:{ce.minute:02d}", "%H:%M")
        window_min = (cwe - cws).total_seconds() / 60
    except Exception:
        window_min = 180
    trig  = getattr(config, "trigger_soc_percent", 40)
    tgt   = getattr(config, "target_soc_percent",  90)
    batt  = getattr(config, "battery_kwh",         210)
    chkw  = getattr(config, "depot_charger_kw",    60)
    cheff = getattr(config, "depot_charger_efficiency", 0.85)
    kwh_needed  = max(0, tgt - trig) / 100 * batt
    charge_min  = kwh_needed / max(0.1, chkw * cheff) * 60
    round_trip  = depot_dead * 2 + charge_min + min_break
    max_conc    = max(1, fleet // 5)

    rows = []
    for _, hw_row in headway_df.iterrows():
        try:
            tf_str  = str(hw_row["time_from"]).strip()
            tt_str  = str(hw_row["time_to"]).strip()
            cfg_hw  = int(hw_row["headway_min"])
        except Exception:
            continue

        travel   = _tt_for_band(tf_str)
        if travel is None:
            continue
        cycle    = travel * 2 + min_break * 2
        nat_gap  = cycle / fleet
        rec_hw   = math.ceil(nat_gap) + 1   # +1 min buffer above floor
        final    = max(cfg_hw, rec_hw)

        if cfg_hw < rec_hw:
            status = "⚠️ Below floor"
            note   = f"Set to {rec_hw} for even {rec_hw}-min spacing"
        elif cfg_hw == rec_hw or cfg_hw == rec_hw - 1:
            status = "✅ Optimal"
            note   = f"Achieves even {cfg_hw}-min spacing"
        else:
            # configured is above floor — check if it's achievable
            status = "✅ Achievable"
            note   = f"Uniform {cfg_hw}-min spacing. Floor={rec_hw-1}"

        rows.append({
            "From":        tf_str,
            "To":          tt_str,
            "Configured":  cfg_hw,
            "Physics Floor": math.ceil(nat_gap),
            "Recommended": rec_hw,
            "Status":      status,
            "Note":        note,
        })

    # Charging impact summary
    buses_absent  = max_conc
    buses_remain  = fleet - buses_absent
    if buses_remain > 0:
        # Use first band's travel time as representative
        first_tt = _tt_for_band(str(headway_df.iloc[0]["time_from"]).strip()) or 40
        chg_cycle = first_tt * 2 + min_break * 2
        chg_hw    = chg_cycle / buses_remain
        rows.append({
            "From":          "P5 window",
            "To":            "(charging)",
            "Configured":    "—",
            "Physics Floor": "—",
            "Recommended":   "—",
            "Status":        "ℹ️ Info",
            "Note":          (
                f"{buses_absent} bus(es) away charging (~{round_trip:.0f} min round-trip). "
                f"{buses_remain} buses remain → expected headway ~{chg_hw:.0f} min. "
                f"Unavoidable without terminal charger."
            ),
        })

    return pd.DataFrame(rows) if rows else None


def _even_spacing_min(config, headway_df, travel_time_df) -> dict:
    """
    Compute the even-spacing minimum headway per band using the coverage formula:
      H_min = ceil((cycle_time + charging_RT) / fleet) + 3 safety buffer

    Also computes k=1.0 and k=1.1 recommendations preserving peak < off-peak ordering.

    Returns dict with keys:
      'by_band': list of {band, even_min, rec_k10, rec_k11}
      'peak_even_min': minimum across peak bands
      'offpeak_even_min': minimum across off-peak bands
      'why': explanation string
      'what_to_do': action string
    """
    import math
    from datetime import datetime as _dt3

    if headway_df is None or headway_df.empty:
        return {}

    fleet     = max(1, config.fleet_size)
    min_break = config.preferred_layover_min

    # Charging RT
    try:
        nodes = [config.start_point, config.end_point]
        nodes += [n.strip() for n in getattr(config, "intermediates", []) if n and n.strip()]
        min_tt_depot = float("inf")
        for node in nodes:
            try:
                tt = config.get_travel_time(config.depot, node)
                min_tt_depot = min(min_tt_depot, tt)
            except Exception:
                pass
        nearest_tt = min_tt_depot if min_tt_depot < float("inf") else 30.0
    except Exception:
        nearest_tt = 30.0

    trig    = getattr(config, "trigger_soc_percent", 40)
    tgt     = getattr(config, "target_soc_percent",  90)
    batt    = getattr(config, "battery_kwh",         210)
    chkw    = getattr(config, "depot_charger_kw",    60)
    cheff   = getattr(config, "depot_charger_efficiency", 0.85)
    min_chg = getattr(config, "min_charge_duration_min", 15)
    kwh     = max(0, tgt - trig) / 100 * batt
    chg_min = max(min_chg, kwh / max(0.1, chkw * cheff) * 60)

    # at_far_loc gate: last ~10% of P5 window may charge from far terminal
    try:
        cs_h = config.p5_charging_start.hour + config.p5_charging_start.minute / 60
        ce_h = config.p5_charging_end.hour   + config.p5_charging_end.minute   / 60
    except Exception:
        cs_h, ce_h = 11.0, 16.0

    def _rt_for_band(time_from_str):
        try:
            h = int(str(time_from_str).split(":")[0])
            band_h = h
        except Exception:
            band_h = 12
        # Last 90-min of P5 window may have at_far_loc bypassed → longer RT
        # Use worst-case far-terminal RT for bands overlapping that window
        p5_near_end = ce_h - 1.5
        if band_h >= p5_near_end:
            # Bus may charge from far terminal: full route time to depot
            try:
                far_loc = config.end_point if config.start_point == config.start_point else config.start_point
                for term in [config.start_point, config.end_point]:
                    try:
                        full_tt = config.get_travel_time(term, config.depot)
                        return full_tt + chg_min + nearest_tt
                    except Exception:
                        pass
            except Exception:
                pass
        return nearest_tt * 2 + chg_min

    def _travel_for_band(time_from_str):
        try:
            t = _dt3.strptime(str(time_from_str).strip(), "%H:%M")
        except Exception:
            return 50.0
        if travel_time_df is not None and not travel_time_df.empty:
            for _, row in travel_time_df.iterrows():
                try:
                    tf = _dt3.strptime(str(row["time_from"]).strip(), "%H:%M")
                    tt = _dt3.strptime(str(row["time_to"]).strip(),   "%H:%M")
                    if tf <= t < tt:
                        return float(row.get("up_min", row.get("dn_min", 50)))
                except Exception:
                    continue
        try:
            return float(config.get_travel_time(config.start_point, config.end_point))
        except Exception:
            return 50.0

    # Identify peak bands (lowest configured headway = tighter service intent)
    try:
        min_hw_configured = int(headway_df["headway_min"].min())
    except Exception:
        min_hw_configured = 15

    band_results = []
    for _, row in headway_df.iterrows():
        try:
            tf_str = str(row["time_from"]).strip()
            tt_str = str(row["time_to"]).strip()
            cfg_hw = int(row["headway_min"])
        except Exception:
            continue
        travel  = _travel_for_band(tf_str)
        cycle   = travel * 2 + min_break * 2
        rt      = _rt_for_band(tf_str)
        even_min = math.ceil((cycle + rt) / fleet) + 3
        is_peak  = (cfg_hw == min_hw_configured)
        band_results.append({
            "band":     f"{tf_str}–{tt_str}",
            "tf":       tf_str,
            "cfg":      cfg_hw,
            "cycle":    cycle,
            "rt":       round(rt, 0),
            "even_min": even_min,
            "is_peak":  is_peak,
        })

    if not band_results:
        return {}

    peak_even    = max((b["even_min"] for b in band_results if b["is_peak"]),  default=27)
    offpeak_even = max((b["even_min"] for b in band_results if not b["is_peak"]), default=28)

    # Enforce ordering: off-peak > peak
    offpeak_even = max(offpeak_even, peak_even + 1)

    # k=1.0 recs
    def _round5(x): return math.ceil(x / 5) * 5
    delta_k10 = max(2, round(0.15 * peak_even))
    peak_k10    = peak_even
    offpeak_k10 = peak_k10 + delta_k10

    # k=1.1 recs
    peak_k11    = math.ceil(1.1 * peak_k10)
    delta_k11   = max(2, round(0.15 * peak_k11))
    offpeak_k11 = peak_k11 + delta_k11

    for b in band_results:
        b["rec_k10"] = peak_k10 if b["is_peak"] else offpeak_k10
        b["rec_k11"] = peak_k11 if b["is_peak"] else offpeak_k11

    # Why / what-to-do strings
    max_gap_approx = round(nearest_tt * 2 + chg_min)
    why = (f"Charging round-trip = {max_gap_approx} min "
           f"({nearest_tt:.0f} min travel × 2 + {chg_min:.0f} min charge). "
           f"No terminal charger at this route's terminus.")
    what = (f"Set peak headway ≥ {peak_even} min and off-peak ≥ {offpeak_even} min, "
            f"OR install a charger within {int(nearest_tt*0.6):.0f} min of the route.")

    return {
        "by_band":        band_results,
        "peak_even_min":  peak_even,
        "offpeak_even_min": offpeak_even,
        "peak_k10":       peak_k10,
        "offpeak_k10":    offpeak_k10,
        "peak_k11":       peak_k11,
        "offpeak_k11":    offpeak_k11,
        "charging_rt":    max_gap_approx,
        "why":            why,
        "what":           what,
    }


def _build_config_excel(config, headway_df, travel_time_df) -> bytes:
    """
    Export the current route config as an Excel file in the same format as
    the input template.  Teal header colour distinguishes dashboard exports.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from datetime import datetime as _dtnow

    TEAL = "0D7377"   # teal = dashboard export marker
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    # ── Sheet 1: Route_Config ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Route_Config"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45

    def _hdr(cell_ref, val):
        c = ws[cell_ref]
        c.value = val
        c.fill  = PatternFill("solid", fgColor=TEAL)
        c.font  = WHITE_FONT

    _hdr("A1", "Parameter")
    _hdr("B1", "Value")
    _hdr("C1", "Notes / Description")

    def _row(row, param, val, note=""):
        ws.cell(row, 1, param)
        ws.cell(row, 2, val)
        ws.cell(row, 3, note)

    r = 2
    _row(r, "route_code",      config.route_code);                                 r+=1
    _row(r, "route_name",      config.route_name);                                 r+=1
    _row(r, "depot",           config.depot);                                      r+=1
    _row(r, "start_point",     config.start_point);                                r+=1
    _row(r, "end_point",       config.end_point);                                  r+=1
    for i, inter in enumerate(getattr(config, "intermediates", []), 1):
        _row(r, f"intermediate_{i}", inter or "");                                 r+=1
    _row(r, "fleet_size",              config.fleet_size);                         r+=1
    _row(r, "battery_kwh",             config.battery_kwh);                       r+=1
    _row(r, "consumption_rate (kWh/km)", config.consumption_rate);                 r+=1
    _row(r, "initial_soc_percent",     config.initial_soc_percent);               r+=1
    _row(r, "avg_speed_kmph",          getattr(config,"avg_speed_kmph",30.0));    r+=1
    _row(r, "depot_charger_kw",        config.depot_charger_kw);                  r+=1
    _row(r, "depot_charger_efficiency",config.depot_charger_efficiency);          r+=1
    _row(r, "terminal_charger_kw",     getattr(config,"terminal_charger_kw",0));  r+=1
    _row(r, "trigger_soc_percent",     config.trigger_soc_percent);               r+=1
    _row(r, "target_soc_percent",      config.target_soc_percent);                r+=1
    _row(r, "min_soc_percent",         config.min_soc_percent);                   r+=1
    _row(r, "min_charge_duration_min", getattr(config,"min_charge_duration_min",15)); r+=1
    _row(r, "midday_charge_soc_percent", getattr(config,"midday_charge_soc_percent",65)); r+=1
    _row(r, "operating_start",         config.operating_start.strftime("%H:%M")); r+=1
    _row(r, "operating_end",           config.operating_end.strftime("%H:%M"));   r+=1
    _row(r, "shift_split",             config.shift_split.strftime("%H:%M"));     r+=1
    _row(r, "preferred_layover_min",   config.preferred_layover_min);             r+=1
    _row(r, "max_layover_min",         getattr(config,"max_layover_min",20));     r+=1
    _row(r, "off_peak_layover_extra_min", getattr(config,"off_peak_layover_extra_min",0)); r+=1
    _row(r, "min_layover_min",         config.min_layover_min);                   r+=1
    _row(r, "dead_run_buffer_min",     config.dead_run_buffer_min);               r+=1
    _row(r, "max_headway_deviation_min", config.max_headway_deviation_min);       r+=1
    _row(r, "max_km_per_bus",          getattr(config,"max_km_per_bus",0));       r+=1
    _row(r, "min_km_per_bus",          getattr(config,"min_km_per_bus",0));       r+=1
    try:
        cs = config.p5_charging_start
        ce = config.p5_charging_end
        _row(r, "p5_charging_start",  cs.strftime("%H:%M") if hasattr(cs,"strftime") else str(cs)); r+=1
        _row(r, "p5_charging_end",    ce.strftime("%H:%M") if hasattr(ce,"strftime") else str(ce)); r+=1
    except Exception:
        pass
    _row(r, "dashboard_export",        _dtnow.now().strftime("%Y-%m-%d %H:%M"),
         "Teal header = exported from eBus Scheduler dashboard");                  r+=1

    # ── Sheet 2: Headway_Profile ──────────────────────────────────────────────
    ws_hw = wb.create_sheet("Headway_Profile")
    for col, hdr in enumerate(["Time From", "Time To", "headway_min", "Notes"], 1):
        c = ws_hw.cell(1, col, hdr)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.font = WHITE_FONT
    for i, (_, row) in enumerate(headway_df.iterrows(), 2):
        ws_hw.cell(i, 1, str(row.get("time_from", "")))
        ws_hw.cell(i, 2, str(row.get("time_to", "")))
        ws_hw.cell(i, 3, int(row.get("headway_min", 20)))

    # ── Sheet 3: TravelTime_Profile ───────────────────────────────────────────
    ws_tt = wb.create_sheet("TravelTime_Profile")
    for col, hdr in enumerate(["Time From", "Time To", "up_min", "dn_min", "Notes"], 1):
        c = ws_tt.cell(1, col, hdr)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.font = WHITE_FONT
    for i, (_, row) in enumerate(travel_time_df.iterrows(), 2):
        ws_tt.cell(i, 1, str(row.get("time_from", "")))
        ws_tt.cell(i, 2, str(row.get("time_to", "")))
        ws_tt.cell(i, 3, row.get("up_min", ""))
        ws_tt.cell(i, 4, row.get("dn_min", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_full_schedule_excel(config, buses) -> bytes:
    """
    Build a rich schedule Excel workbook for download.
    Two sheets:
      1. Trip Schedule — every trip with sequential numbering, break times,
         running distance, SOC curve, and classification columns
      2. Bus Summary — per-bus totals
    Returns raw bytes ready for st.download_button.
    """
    import io
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_trips = wb.active
    ws_trips.title = "Trip Schedule"

    # Headers
    headers = [
        "Trip #", "Bus", "Type", "Revenue?", "Direction",
        "From", "To", "Departure", "Arrival",
        "Travel (min)", "Break After (min)",
        "Distance (km)", "Cumulative KM",
        "SOC Before (%)", "SOC After (%)",
        "Shift", "Notes"
    ]
    header_fill   = PatternFill("solid", fgColor="0D5CA8")   # teal-blue = dashboard export
    header_font   = Font(color="FFFFFF", bold=True)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin          = Side(style="thin", color="CCCCCC")
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws_trips.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        cell.border = cell_border

    ws_trips.row_dimensions[1].height = 28

    # Revenue trip counter per bus (sequential, excluding depot trips)
    trip_seq = 0
    row_idx  = 2
    for bus in buses:
        soc         = config.initial_soc_percent
        cum_km      = 0.0
        trip_number = 0   # sequential within this bus (all trips)
        rev_number  = 0   # revenue-only counter

        for i, trip in enumerate(bus.trips):
            # SOC before this trip
            soc_before = round(soc, 1)
            soc -= (trip.distance_km * config.consumption_rate / config.battery_kwh) * 100
            if trip.trip_type == "Charging":
                soc = min(100.0, soc + config.depot_flow_rate_kw *
                          (trip.travel_time_min / 60) / config.battery_kwh * 100)
            soc_after = round(soc, 1)
            cum_km   += trip.distance_km

            trip_number += 1
            is_revenue   = trip.trip_type == "Revenue"
            is_shuttle   = trip.trip_type == "Shuttle"
            is_dead      = trip.trip_type == "Dead"
            is_charging  = trip.trip_type == "Charging"

            if is_revenue or is_shuttle:
                rev_number += 1
                seq_label   = str(rev_number)
            else:
                seq_label   = "—"

            # Break after this trip
            brk = _true_driver_break(bus, i) or 0

            revenue_label = (
                "✓ Revenue"  if is_revenue  else
                "Shuttle"    if is_shuttle  else
                "Dead Run"   if is_dead     else
                "Charging"
            )
            direction_label = (
                trip.direction if trip.direction not in ("DEPOT", None, "") else "—"
            )
            notes = ""
            if is_charging:
                notes = f"Depot charge {trip.travel_time_min} min"
            elif is_dead:
                notes = "Non-revenue positioning"
            elif is_shuttle:
                notes = "Shuttle (not counted in revenue KPIs)"

            row_data = [
                seq_label,
                trip.assigned_bus or bus.bus_id,
                trip.trip_type,
                revenue_label,
                direction_label,
                trip.start_location,
                trip.end_location,
                trip.actual_departure.strftime("%H:%M") if trip.actual_departure else "",
                trip.actual_arrival.strftime("%H:%M") if trip.actual_arrival else "",
                trip.travel_time_min,
                brk if brk > 0 else "",
                round(trip.distance_km, 2),
                round(cum_km, 2),
                soc_before,
                soc_after,
                trip.shift,
                notes,
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws_trips.cell(row=row_idx, column=col, value=val)
                cell.border    = cell_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Colour code by trip type
                if is_revenue:
                    cell.fill = PatternFill("solid", fgColor="E8F5E9")   # light green
                elif is_charging:
                    cell.fill = PatternFill("solid", fgColor="FFF8E1")   # light amber
                elif is_dead:
                    cell.fill = PatternFill("solid", fgColor="F5F5F5")   # light grey
            row_idx += 1

    # Column widths
    col_widths = [8, 12, 12, 12, 12, 20, 20, 10, 10, 13, 16, 14, 14, 13, 13, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws_trips.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: Bus Summary
    ws_bus = wb.create_sheet("Bus Summary")
    bus_headers = [
        "Bus", "Revenue Trips", "Shuttle Trips", "Dead Runs", "Charging Stops",
        "Total KM", "Revenue KM", "Dead KM",
        "Final SOC (%)", "First Departure", "Last Arrival", "Total Active (min)"
    ]
    for col, h in enumerate(bus_headers, 1):
        cell = ws_bus.cell(row=1, column=col, value=h)
        cell.fill  = PatternFill("solid", fgColor="0D5CA8")
        cell.font  = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r_idx, bus in enumerate(buses, 2):
        rev   = [t for t in bus.trips if t.trip_type == "Revenue"]
        shut  = [t for t in bus.trips if t.trip_type == "Shuttle"]
        dead  = [t for t in bus.trips if t.trip_type == "Dead"]
        chg   = [t for t in bus.trips if t.trip_type == "Charging"]
        first = next((t.actual_departure for t in bus.trips if t.actual_departure), None)
        last  = next((t.actual_arrival   for t in reversed(bus.trips) if t.actual_arrival), None)
        active_min = ((last - first).total_seconds() / 60) if first and last else 0
        row_vals = [
            bus.bus_id, len(rev), len(shut), len(dead), len(chg),
            round(bus.total_km, 1),
            round(sum(t.distance_km for t in rev), 1),
            round(sum(t.distance_km for t in dead), 1),
            round(bus.soc_percent, 1),
            first.strftime("%H:%M") if first else "",
            last.strftime("%H:%M")  if last  else "",
            round(active_min, 0),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws_bus.cell(row=r_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")

    # Dashboard-export marker in a metadata sheet
    ws_meta = wb.create_sheet("Export_Info")
    from datetime import datetime as _dt_now
    ws_meta["A1"] = "Generated by"
    ws_meta["B1"] = "eBus Scheduler Dashboard Export"
    ws_meta["A2"] = "Route"
    ws_meta["B2"] = config.route_code
    ws_meta["A3"] = "Route Name"
    ws_meta["B3"] = config.route_name
    ws_meta["A4"] = "Exported at"
    ws_meta["B4"] = _dt_now.now().strftime("%Y-%m-%d %H:%M")
    ws_meta["A5"] = "Note"
    ws_meta["B5"] = "Teal header colour identifies dashboard-exported config files"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _apply_config_overrides(config, overrides):
    from src.models import RouteConfig
    return RouteConfig(
        route_code=config.route_code, route_name=config.route_name,
        depot=config.depot, start_point=config.start_point,
        end_point=config.end_point, intermediates=config.intermediates,
        fleet_size=overrides.get("fleet_size", config.fleet_size),
        battery_kwh=overrides.get("battery_kwh", config.battery_kwh),
        consumption_rate=overrides.get("consumption_rate", config.consumption_rate),
        initial_soc_percent=overrides.get("initial_soc_percent", config.initial_soc_percent),
        depot_charger_kw=overrides.get("depot_charger_kw", config.depot_charger_kw),
        depot_charger_efficiency=overrides.get("depot_charger_efficiency", config.depot_charger_efficiency),
        terminal_charger_kw=config.terminal_charger_kw,
        terminal_charger_efficiency=config.terminal_charger_efficiency,
        trigger_soc_percent=overrides.get("trigger_soc_percent", config.trigger_soc_percent),
        target_soc_percent=overrides.get("target_soc_percent", config.target_soc_percent),
        min_soc_percent=overrides.get("min_soc_percent", config.min_soc_percent),
        min_charge_duration_min=overrides.get("min_charge_duration_min", config.min_charge_duration_min),
        operating_start=overrides.get("operating_start", config.operating_start),
        operating_end=overrides.get("operating_end", config.operating_end),
        shift_split=overrides.get("shift_split", config.shift_split),
        min_layover_min=overrides.get("min_layover_min", config.min_layover_min),
        preferred_layover_min=overrides.get("preferred_layover_min", config.preferred_layover_min),
        dead_run_buffer_min=overrides.get("dead_run_buffer_min", config.dead_run_buffer_min),
        max_headway_deviation_min=overrides.get("max_headway_deviation_min", config.max_headway_deviation_min),
        km_balance_tolerance_pct=overrides.get("km_balance_tolerance_pct", config.km_balance_tolerance_pct),
        segment_distances=config.segment_distances, segment_times=config.segment_times,
        location_coords=config.location_coords,
        min_km_per_bus=overrides.get("min_km_per_bus", config.min_km_per_bus),
        max_km_per_bus=overrides.get("max_km_per_bus", getattr(config, 'max_km_per_bus', 0)),
        max_layover_min=overrides.get("max_layover_min",
                                      getattr(config, "max_layover_min", 20)),
        midday_charge_soc_percent=overrides.get("midday_charge_soc_percent",
                                                 getattr(config, "midday_charge_soc_percent", 65.0)),
        off_peak_layover_extra_min=overrides.get("off_peak_layover_extra_min",
                                                  getattr(config, "off_peak_layover_extra_min", 0)),
        avg_speed_kmph=overrides.get("avg_speed_kmph",
                                     getattr(config, "avg_speed_kmph", 30.0)),
    )

def _run_core(config, headway_df, travel_time_df, optimize,
              scheduling_mode: str = "planning"):
    enrich_distances(config)
    trips = generate_trips(config, headway_df, travel_time_df,
                           scheduling_mode=scheduling_mode)
    revenue_trips = [t for t in trips if t.trip_type == "Revenue"]
    if optimize:
        from src.optimizer import optimize_schedule
        buses, metrics, _ = optimize_schedule(config, headway_df, travel_time_df,
                                              verbose=False,
                                              scheduling_mode=scheduling_mode)
        assigned_rev = sum(1 for b in buses for t in b.trips if t.trip_type == 'Revenue')
        metrics = compute_metrics(config, buses,
                                  total_revenue_trips=len(revenue_trips),
                                  assigned_revenue_trips=assigned_rev)
    else:
        buses = schedule_buses(config, trips,
                               headway_df=headway_df, travel_time_df=travel_time_df,
                               scheduling_mode=scheduling_mode)
        # Bus-driven scheduler creates new Trip objects (not pool references).
        # Count Revenue trips directly from bus schedules.
        assigned_rev = sum(1 for b in buses for t in b.trips if t.trip_type == 'Revenue')
        metrics = compute_metrics(config, buses,
                                  total_revenue_trips=len(revenue_trips),
                                  assigned_revenue_trips=assigned_rev)
    compliance = check_compliance(config, buses, headway_df=headway_df)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        write_schedule(config, buses, f.name)
        out = Path(f.name).read_bytes()
    return config, buses, metrics, trips, out, compliance

def auto_detect_fleet(raw_config, headway_df, travel_time_df, max_fleet=20):
    """
    Find the minimum fleet_size (1..max_fleet) where all P1-P6 rules PASS
    and every revenue trip is assigned.  Distances are enriched once and
    shared across all trials via the same dict reference.
    Returns (detected_n, config, buses, metrics, trips, output_bytes, compliance).
    """
    # Enrich distances once (mutations are in-place on the shared dict)
    seed_cfg = _apply_config_overrides(raw_config, {"fleet_size": 1})
    enrich_distances(seed_cfg)

    last_result = None
    for n in range(1, max_fleet + 1):
        cfg = _apply_config_overrides(raw_config, {"fleet_size": n})
        trips = generate_trips(cfg, headway_df, travel_time_df,
                               scheduling_mode="planning")
        buses = schedule_buses(cfg, trips,
                               headway_df=headway_df,
                               travel_time_df=travel_time_df,
                               scheduling_mode="planning")
        rev_total    = sum(1 for t in trips   if t.trip_type == "Revenue")
        rev_assigned = sum(1 for b in buses for t in b.trips if t.trip_type == "Revenue")
        compliance   = check_compliance(cfg, buses, headway_df=headway_df)
        hard_pass    = all(r["status"] == "PASS"
                           for r in compliance if r.get("priority", 99) <= 6)
        if hard_pass and rev_assigned == rev_total:
            metrics = compute_metrics(cfg, buses, total_revenue_trips=rev_total)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                write_schedule(cfg, buses, f.name)
                out = Path(f.name).read_bytes()
            return n, cfg, buses, metrics, trips, out, compliance
        # keep last attempt as fallback
        last_result = (n, cfg, buses, rev_total, compliance)

    # Fallback: return best attempt with max fleet
    n, cfg, buses, rev_total, compliance = last_result
    metrics = compute_metrics(cfg, buses, total_revenue_trips=rev_total)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        write_schedule(cfg, buses, f.name)
        out = Path(f.name).read_bytes()
    return n, cfg, buses, metrics, [], out, compliance


def run_pipeline(uploaded_file, optimize, config_overrides=None, headway_overrides=None, service_max=False):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(uploaded_file.getvalue()); tmp_path = tmp.name
    config, headway_df, travel_time_df = load_config(tmp_path)
    st.session_state["raw_config"] = config
    st.session_state["raw_headway_df"] = headway_df.copy()
    st.session_state["raw_travel_time_df"] = travel_time_df.copy()
    if config_overrides: config = _apply_config_overrides(config, config_overrides)
    if headway_overrides is not None: headway_df = headway_overrides

    # Auto fleet-size detection when fleet_size == 0 in the Excel
    if config.fleet_size == 0:
        st.session_state["auto_fleet_mode"] = True
        detected_n, cfg, buses, metrics, trips, out, compliance = auto_detect_fleet(
            config, headway_df, travel_time_df)
        st.session_state["detected_fleet_size"] = detected_n
        return cfg, buses, metrics, trips, out, compliance

    st.session_state["auto_fleet_mode"] = False
    st.session_state.pop("detected_fleet_size", None)
    if service_max:
        from src.city_models import RouteInput
        ri = RouteInput(config=config, headway_df=headway_df, travel_time_df=travel_time_df)
        nat_hw = _natural_headway(ri) if _CITY_OK else None
        if nat_hw:
            headway_df = _flat_headway_df(ri, nat_hw)
            st.session_state["service_max_headway"] = nat_hw
    # Derive scheduling_mode from UI flags:
    #   service_max → efficiency (flat headway, physics-derived spacing)
    #   optimize    → efficiency (KPI-driven, throughput-focused)
    #   otherwise   → planning  (strict headway enforcement)
    _sched_mode = "efficiency" if (optimize or service_max) else "planning"
    return _run_core(config, headway_df, travel_time_df, optimize,
                     scheduling_mode=_sched_mode)

def rerun_from_overrides(config_overrides, headway_overrides=None, optimize=False, service_max=False):
    raw_config = st.session_state.get("raw_config")
    raw_hw = st.session_state.get("raw_headway_df")
    raw_tt = st.session_state.get("raw_travel_time_df")
    if raw_config is None: return None
    config = _apply_config_overrides(raw_config, config_overrides)
    headway_df = headway_overrides if headway_overrides is not None else raw_hw
    # Honour auto-detect if user set fleet_size back to 0 in the Config tab
    if config.fleet_size == 0:
        st.session_state["auto_fleet_mode"] = True
        detected_n, cfg, buses, metrics, trips, out, compliance = auto_detect_fleet(
            config, headway_df, raw_tt)
        st.session_state["detected_fleet_size"] = detected_n
        return cfg, buses, metrics, trips, out, compliance
    st.session_state["auto_fleet_mode"] = False
    st.session_state.pop("detected_fleet_size", None)
    if service_max and _CITY_OK:
        from src.city_models import RouteInput
        ri2 = RouteInput(config=config, headway_df=headway_df, travel_time_df=raw_tt)
        nat_hw2 = _natural_headway(ri2)
        headway_df = _flat_headway_df(ri2, nat_hw2)
        st.session_state["service_max_headway"] = nat_hw2
    _sched_mode = "efficiency" if (optimize or service_max) else "planning"
    return _run_core(config, headway_df, raw_tt, optimize,
                     scheduling_mode=_sched_mode)


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🚌 eBus Scheduler")

    if _CITY_OK:
        app_mode = st.radio(
            "Mode", ["🚌 Single Route", "🏙️ Citywide"], index=0,
            help="Single Route: one config. Citywide: multiple configs with fleet rebalancing.",
        )
    else:
        app_mode = "🚌 Single Route"

    st.divider()

    if app_mode == "🚌 Single Route":
        st.caption("Upload route config Excel to generate a schedule.")
        uploaded = st.file_uploader("Config Excel", type=["xlsx"], label_visibility="collapsed")
        single_mode = st.radio(
            "Scheduling mode",
            ["📋 Planning-Compliant", "⚡ Efficiency-Maximising", "🎯 Service Maximization"],
            index=0, label_visibility="collapsed",
        )
        if single_mode == "📋 Planning-Compliant":
            optimize = False; service_max_single = False
            st.info("📋 **Planning-Compliant** — follows headway profile strictly. "
                    "Fleet size from config.", icon="📋")
        elif single_mode == "⚡ Efficiency-Maximising":
            optimize = True; service_max_single = False
            st.info("⚡ **Efficiency-Maximising** — finds minimum fleet satisfying all rules. "
                    "Config fleet size overridden; headway floor still enforced.", icon="⚡")
        else:
            optimize = False; service_max_single = True
            st.info("🎯 **Service Maximization** — uses config fleet, ignores headway profile. "
                    "Computes minimum achievable constant headway for even spacing.", icon="🎯")
        run_btn = st.button("▶ Generate Schedule", type="primary", disabled=uploaded is None)
        st.divider()
        st.caption("Rules enforced: P4 break from config, P2 via nearest node, P5 midday charge, P3 SOC ≥ 20%.")
    else:
        st.caption("Upload config Excel files for all routes.")
        uploaded_files = st.file_uploader(
            "Route Config Files", type=["xlsx", "xlsm"],
            accept_multiple_files=True, label_visibility="collapsed",
        )
        city_sched_mode = st.radio(
            "Scheduling mode",
            ["📋 Planning-Compliant", "⚡ Efficiency-Maximising", "🎯 Service Maximization"],
            index=0, label_visibility="collapsed",
        )
        if city_sched_mode == "📋 Planning-Compliant":
            city_mode = "planning"
            st.info("📋 **Planning-Compliant** — headway profile respected, surplus rebalanced.", icon="📋")
        elif city_sched_mode == "⚡ Efficiency-Maximising":
            city_mode = "efficiency"
            st.info("⚡ **Efficiency-Maximising** — minimum fleet per route, KPI-driven. "
                    "Config fleet size overridden; headway floor still enforced.", icon="⚡")
        else:
            city_mode = "service_max"
            st.info("🎯 **Service Maximization** — config fleet used as-is. "
                    "Headway profile ignored; constant minimum-achievable headway computed per route.", icon="🎯")
        total_fleet_override = st.number_input(
            "Total Fleet Override", min_value=0, value=0, step=1,
            help="0 = use sum from configs. >0 = cap total citywide fleet.",
        )
        city_run_btn = st.button(
            "▶ Generate Citywide Schedule", type="primary",
            disabled=not uploaded_files,
        )
        st.divider()
        st.caption("📋 Planning-Compliant · ⚡ Efficiency-Maximising · 🎯 Service Maximization")


# ── Main ──────────────────────────────────────────────────────────────────────
# Ensure variables exist regardless of which sidebar mode was active
if 'uploaded' not in dir(): uploaded = None
if 'optimize' not in dir(): optimize = False
if 'run_btn' not in dir(): run_btn = False
if 'uploaded_files' not in dir(): uploaded_files = []
if 'city_run_btn' not in dir(): city_run_btn = False
if 'total_fleet_override' not in dir(): total_fleet_override = 0
if 'service_max_single' not in dir(): service_max_single = False
if 'city_mode' not in dir(): city_mode = "planning"
# Legacy compat: city_optimize used in Fleet Config re-run button
city_optimize = (city_mode == "efficiency")

if app_mode == "🚌 Single Route":
    # ═══════════════════════════════ SINGLE ROUTE ════════════════════════════════

    if uploaded is None:
        st.markdown("## Welcome to eBus Scheduler")
        c1, c2, c3, c4 = st.columns(4)
        icons = ["📁", "⚙️", "📊", "⬇️"]
        steps = [("Upload", "Drop your config Excel with route, fleet, headway, and travel time data."),
                 ("Generate", "Engine creates trips, assigns buses with min-break enforcement and midday charging."),
                 ("Review", "Check compliance, route depiction, headways, SOC curves, and KM balance."),
                 ("Download", "Get formatted schedule Excel with full trip-by-trip detail.")]
        for col, (icon, (title, desc)) in zip([c1, c2, c3, c4], zip(icons, steps)):
            with col:
                st.markdown(f"### {icon} {title}")
                st.caption(desc)

    elif run_btn:
        with st.spinner("Running scheduler..."):
            try:
                result = run_pipeline(uploaded, optimize, service_max=service_max_single)
            except ConfigError as e:
                st.error(f"Config error: {e}"); st.stop()
            except Exception as e:
                st.error(f"Error: {e}"); st.stop()
        config, buses, metrics, trips, output_bytes, compliance = result
        st.session_state.update({
            "config": config, "buses": buses, "metrics": metrics, "trips": trips,
            "output_bytes": output_bytes, "compliance": compliance,
            "prev_compliance": None, "has_results": True,
        })

    if st.session_state.get("has_results"):
        config = st.session_state["config"]
        buses  = st.session_state["buses"]
        metrics = st.session_state["metrics"]
        output_bytes = st.session_state["output_bytes"]
        compliance = st.session_state.get("compliance", [])
        prev_compliance = st.session_state.get("prev_compliance")

        # ── Route title + download ────────────────────────────────────────────
        col_title, col_dl = st.columns([3, 1])
        with col_title:
            st.markdown(f"## Route {config.route_code} — {config.route_name}")
        with col_dl:
            st.download_button(
                f"⬇ Download Schedule",
                data=output_bytes,
                file_name=f"{config.route_code}_schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── KPI bar ──────────────────────────────────────────────────────────
        all_pass = all(r["status"] in ("PASS",) for r in compliance if r.get("priority", 99) <= 6)
        fail_count = sum(1 for r in compliance if r["status"] == "FAIL")
        soc_ok = metrics.min_soc_seen >= 25
        km_ok = metrics.km_range <= 20
        shuttle_count = sum(1 for b in buses for t in b.trips if t.trip_type == "Shuttle")
        # trip_ok: all pool trips are covered (revenue + shuttle together)
        # Shuttle trips are passenger-carrying corridor legs added by the scheduler.
        # They are not in the pool, so we only check revenue against pool target.
        trip_ok = metrics.revenue_trips_assigned >= metrics.revenue_trips_total
        unserved = max(0, metrics.revenue_trips_total - metrics.revenue_trips_assigned)

        # Revenue Trips KPI: show R · S · D counts on one line
        # R = full Revenue trips (pool), S = Shuttle trips (corridor legs), D = Dead runs
        dead_count = sum(1 for b in buses for t in b.trips if t.trip_type == "Dead")
        rev_parts  = [f"{metrics.revenue_trips_assigned}R"]
        if shuttle_count > 0:
            rev_parts.append(f"{shuttle_count}S")
        rev_parts.append(f"{dead_count}D")
        rev_label  = " · ".join(rev_parts)
        total_trips = metrics.revenue_trips_assigned + shuttle_count + dead_count
        rev_sub     = f"{total_trips} total trips · of {metrics.revenue_trips_total} planned revenue"
        rev_status = "ok" if unserved == 0 else "warn" if unserved <= 5 else "bad"

        st.markdown(
            '<div class="kpi-grid">' +
            kpi("Revenue Trips", rev_label, rev_status, sub=rev_sub) +
            kpi("Total KM", f"{metrics.total_km:.0f}") +
            kpi("Dead KM %", f"{metrics.dead_km_ratio:.1%}",
                "ok" if metrics.dead_km_ratio < 0.15 else "warn") +
            kpi("KM Deviation", f"{metrics.km_range:.1f} km",
                "ok" if km_ok else "warn") +
            kpi("Min SOC", f"{metrics.min_soc_seen:.1f}%",
                "ok" if soc_ok else "warn") +
            kpi("Charging Stops", str(metrics.charging_stops)) +
            kpi("Compliance", f"{10-fail_count}/10",
                "ok" if fail_count == 0 else "warn" if fail_count <= 2 else "bad") +
            kpi("Fleet", f"{config.fleet_size} buses") +
            '</div>', unsafe_allow_html=True,
        )

        # ── Tabs ─────────────────────────────────────────────────────────────
        tab_compliance, tab_route, tab_fleet, tab_schedule, tab_bus, tab_config = st.tabs([
            "✅ Compliance", "🗺 Route Depiction", "📊 Fleet Summary",
            "📋 Full Schedule", "🚌 Bus Detail", "⚙️ Config",
        ])

        # ════════════════════════════════════════════════════════════════════
        # TAB 0: Compliance
        # ════════════════════════════════════════════════════════════════════
        with tab_compliance:
            if prev_compliance:
                st.info("Showing before/after comparison from config edit.")

            STATUS = {"PASS": ("✅", "pass"), "FAIL": ("❌", "fail"),
                      "WARN": ("⚠️", "warn"), "INFO": ("✅", "pass")}

            p_rules = [r for r in compliance if r.get("priority", 99) <= 6]
            o_rules = [r for r in compliance if r.get("priority", 99) > 6]

            for section_label, rules, tag_cls, tag_text in [
                ("Priority Rules (P1–P6)", p_rules, "tag-p", "P"),
                ("Operational Rules (O1–O4)", o_rules, "tag-o", "O"),
            ]:
                st.markdown(f'<div class="section-title">{section_label}</div>', unsafe_allow_html=True)
                for i, rule in enumerate(rules):
                    effective = rule["status"]
                    if effective == "INFO" and not rule.get("violations"): effective = "PASS"
                    icon, cls = STATUS.get(effective, ("❓", ""))
                    viols = rule.get("violations", [])
                    if not viols and effective == "WARN": viols = [rule.get("details", "")]

                    # before/after badge
                    badge = ""
                    if prev_compliance:
                        prev_idx = next((j for j, r in enumerate(prev_compliance)
                                         if r["rule"] == rule["rule"]), None)
                        if prev_idx is not None and prev_compliance[prev_idx]["status"] != rule["status"]:
                            pi, _ = STATUS.get(prev_compliance[prev_idx]["status"], ("❓",""))
                            badge = f' <span style="font-size:.75rem;color:#888">{pi}→{icon}</span>'

                    label = f"{icon} {rule['rule']}"
                    with st.expander(label, expanded=(effective == "FAIL")):
                        st.caption(rule.get("details", ""))
                        if viols:
                            for v in viols: st.markdown(f"- {v}")
                        else:
                            st.markdown("_No violations._")

            if prev_compliance:
                changed = [r["rule"] for r in compliance
                           if any(p["rule"] == r["rule"] and p["status"] != r["status"]
                                  for p in prev_compliance)]
                if changed: st.warning(f"Rules changed: {', '.join(changed)}")
                else: st.success("No rule statuses changed.")

        # ════════════════════════════════════════════════════════════════════
        # TAB 1: Route Depiction
        # ════════════════════════════════════════════════════════════════════
        with tab_route:
            # ── Bus filter + diagram ─────────────────────────────────────────
            if not _PLOTLY_OK:
                st.warning("Install **plotly** to enable the route diagram (`pip install plotly` or add to requirements.txt).")
            else:
                filter_col, spacer = st.columns([2, 5])
                with filter_col:
                    bus_opts = ["All buses"] + config.bus_ids()
                    sel = st.selectbox("Filter diagram by bus", bus_opts,
                                       label_visibility="collapsed")
                selected_bus_filter = None if sel == "All buses" else sel
                try:
                    route_fig = build_route_diagram(config, buses,
                                                    selected_bus=selected_bus_filter)
                    st.plotly_chart(route_fig, use_container_width=True,
                                    config={"displayModeBar": True,
                                            "modeBarButtonsToRemove": ["lasso2d","select2d"],
                                            "toImageButtonOptions": {"filename": f"route_{config.route_code}"}})
                except Exception as e:
                    st.warning(f"Could not render route diagram: {e}")
                st.caption(
                    "Straight-line representation: distances in Panel A are cumulative along the route; "
                    "depot is shown off-line for clarity. "
                    "Panel B is a time–space diagram: parallel diagonal lines = even headways; "
                    "converging lines = bunching; gaps = charging or off-peak window."
                )
            st.divider()

            # ── Per-bus trip tables ──────────────────────────────────────────
            st.markdown('<div class="section-title">Per-Bus Trip Detail</div>', unsafe_allow_html=True)
            route_df = build_route_depiction(config, buses)
            for bus in buses:
                bus_data = route_df[route_df["Bus"] == bus.bus_id].reset_index(drop=True)
                if bus_data.empty: continue

                rev = [t for t in bus.trips if t.trip_type == "Revenue"]
                dead_km = sum(t.distance_km for t in bus.trips if t.trip_type == "Dead")
                chg = [t for t in bus.trips if t.trip_type == "Charging"]
                first = next((t.actual_departure for t in bus.trips if t.actual_departure), None)
                last  = next((t.actual_arrival for t in reversed(bus.trips) if t.actual_arrival), None)

                st.markdown(
                    f'<div class="bus-header">'
                    f'<span class="bus-pill">{bus.bus_id}</span>'
                    f'{len(rev)} revenue trips &nbsp;·&nbsp; {bus.total_km:.1f} km &nbsp;·&nbsp;'
                    f' SOC {bus.soc_percent:.1f}% final &nbsp;·&nbsp;'
                    f' {first.strftime("%H:%M") if first else "?"} – {last.strftime("%H:%M") if last else "?"}'
                    f'</div>', unsafe_allow_html=True,
                )

                display_rows = []
                for _, row in bus_data.iterrows():
                    t = row["_type"]; d = row["_dir"]
                    marker = "🔋" if t=="Charging" else "⚫" if t=="Dead" else "🔵" if d=="UP" else "🟢"
                    brk = row["Break"]
                    display_rows.append({
                        "": marker, "Dep": row["Dep"], "Arr": row["Arr"],
                        "From": row["From"], "To": row["To"], "Type": row["Type"],
                        "Dist": row["Dist"],
                        "SOC": f'{row["SOC"]}{"⚠️" if row["SOC"] < 30 else ""}',
                        "Break": f"{brk} min" if brk > 0 else "",
                        "Dead KM": row["Dist"] if t == "Dead" else "",
                    })

                df_show = pd.DataFrame(display_rows)
                st.dataframe(df_show, hide_index=True,
                             height=min(420, 36 * len(df_show) + 38),
                             column_config={"": st.column_config.TextColumn(width="small")})
                st.caption(
                    f"Dead KM: {dead_km:.1f} &nbsp;|&nbsp; "
                    f"Charging: {len(chg)} stop{'s' if len(chg)!=1 else ''} &nbsp;|&nbsp; "
                    f"Shift split: {config.shift_split}"
                )
                st.divider()

        # ════════════════════════════════════════════════════════════════════
        # TAB 2: Fleet Summary
        # ════════════════════════════════════════════════════════════════════
        with tab_fleet:
            fleet_df = build_fleet_df(config, buses)
            st.dataframe(fleet_df, hide_index=True, use_container_width=True)

            col1, col2 = st.columns(2)
            with col1:
                st.markdown('<div class="section-title">KM per Bus</div>', unsafe_allow_html=True)
                km_chart = fleet_df[["Bus","Revenue KM","Dead KM"]].set_index("Bus")
                st.bar_chart(km_chart, color=["#4f46e5","#94a3b8"])
            with col2:
                st.markdown('<div class="section-title">Final SOC per Bus (%)</div>', unsafe_allow_html=True)
                soc_chart = fleet_df[["Bus","Final SOC (%)"]].set_index("Bus")
                st.bar_chart(soc_chart, color="#16a34a")

            st.markdown('<div class="section-title">Departure Headways</div>', unsafe_allow_html=True)
            st.caption("Time gaps between consecutive departures in the same direction. "
                       "Irregular gaps are expected when bus cycle time ≠ fleet × headway — "
                       "see Config tab to tune.")
            hw_data = build_headway_chart_data(config, buses)
            _hw_df_sr = st.session_state.get("raw_headway_df")
            if not hw_data.empty:
                col_up, col_dn = st.columns(2)
                for col, direction, color in [(col_up, "UP", "#4f46e5"), (col_dn, "DN", "#16a34a")]:
                    with col:
                        dir_data = hw_data[hw_data["Direction"]==direction].sort_values("Departure")
                        deps = dir_data["Departure"].tolist()
                        _fig = build_headway_fig(deps, _hw_df_sr, direction, color,
                                                  config.route_code)
                        if _fig:
                            st.plotly_chart(_fig, use_container_width=True)
                        else:
                            st.caption(f"**{direction}** — fewer than 2 departures.")

            # ── Headway stability metrics ─────────────────────────────────────
            st.markdown('<div class="section-title">Schedule Stability</div>',
                        unsafe_allow_html=True)
            st.caption(
                "**On-time %** = departures within ±3 min of ideal uniform spacing. "
                "**Avg drift** = mean deviation from ideal departure (lower = more reliable). "
                "**Headway std** = standard deviation of all gaps (0 = perfectly uniform). "
                "**Stability Score** = 1 / (1 + headway_std): 1.0 = perfect, < 0.5 = poor."
            )
            _stab_c1, _stab_c2, _stab_c3, _stab_c4, _stab_c5 = st.columns(5)
            _ot = getattr(metrics, 'pct_trips_on_time', 0.0)
            _drift_avg = getattr(metrics, 'avg_drift_min', 0.0)
            _drift_max = getattr(metrics, 'max_drift_min', 0.0)
            _hw_std = getattr(metrics, 'headway_std_min', 0.0)
            _stab_score = 1.0 / (1.0 + _hw_std)   # 1.0 = perfect uniform; drops as std grows
            with _stab_c1:
                _ot_status = "ok" if _ot >= 80 else ("warn" if _ot >= 60 else "fail")
                st.markdown(
                    f'<div class="kpi-card kpi-{_ot_status}">'
                    f'<div class="kpi-label">On-Time Departures</div>'
                    f'<div class="kpi-value">{_ot:.0f}%</div>'
                    f'<div class="kpi-sub">±3 min tolerance</div></div>',
                    unsafe_allow_html=True)
            with _stab_c2:
                _std_status = "ok" if _hw_std <= 3 else ("warn" if _hw_std <= 8 else "fail")
                st.markdown(
                    f'<div class="kpi-card kpi-{_std_status}">'
                    f'<div class="kpi-label">Headway Std Dev</div>'
                    f'<div class="kpi-value">{_hw_std:.1f} min</div>'
                    f'<div class="kpi-sub">0 = perfectly uniform</div></div>',
                    unsafe_allow_html=True)
            with _stab_c3:
                _davg_status = "ok" if _drift_avg <= 3 else ("warn" if _drift_avg <= 8 else "fail")
                st.markdown(
                    f'<div class="kpi-card kpi-{_davg_status}">'
                    f'<div class="kpi-label">Avg Schedule Drift</div>'
                    f'<div class="kpi-value">{_drift_avg:.1f} min</div>'
                    f'<div class="kpi-sub">actual vs ideal departure</div></div>',
                    unsafe_allow_html=True)
            with _stab_c4:
                _dmax_status = "ok" if _drift_max <= 5 else ("warn" if _drift_max <= 15 else "fail")
                st.markdown(
                    f'<div class="kpi-card kpi-{_dmax_status}">'
                    f'<div class="kpi-label">Max Schedule Drift</div>'
                    f'<div class="kpi-value">{_drift_max:.1f} min</div>'
                    f'<div class="kpi-sub">worst trip deviation</div></div>',
                    unsafe_allow_html=True)
            with _stab_c5:
                _ss_status = "ok" if _stab_score >= 0.5 else ("warn" if _stab_score >= 0.25 else "fail")
                st.markdown(
                    f'<div class="kpi-card kpi-{_ss_status}">'
                    f'<div class="kpi-label">Stability Score</div>'
                    f'<div class="kpi-value">{_stab_score:.2f}</div>'
                    f'<div class="kpi-sub">1/(1+std) · 1.0 = perfect</div></div>',
                    unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════════
        # TAB 3: Full Schedule
        # ════════════════════════════════════════════════════════════════════
        with tab_schedule:
            schedule_df = build_schedule_df(config, buses)
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                type_filter = st.multiselect("Trip Type", ["Revenue","Dead","Charging"],
                                             default=["Revenue","Dead","Charging"])
            with fc2:
                bus_filter = st.multiselect("Bus", config.bus_ids(), default=config.bus_ids())
            with fc3:
                dir_filter = st.multiselect("Direction", schedule_df["Direction"].unique().tolist(),
                                            default=schedule_df["Direction"].unique().tolist())
            filtered = schedule_df[
                schedule_df["Type"].isin(type_filter) &
                schedule_df["Bus"].isin(bus_filter) &
                schedule_df["Direction"].isin(dir_filter)
            ]
            st.dataframe(filtered, hide_index=True, height=520, use_container_width=True)
            st.caption(f"Showing {len(filtered)} of {len(schedule_df)} trips")

        # ════════════════════════════════════════════════════════════════════
        # TAB 4: Bus Detail
        # ════════════════════════════════════════════════════════════════════
        with tab_bus:
            selected_bus = st.selectbox("Select Bus", config.bus_ids(), label_visibility="collapsed")
            bus_obj = next(b for b in buses if b.bus_id == selected_bus)
            bus_df = build_schedule_df(config, [bus_obj])
            rev_count = len([t for t in bus_obj.trips if t.trip_type=="Revenue"])
            dead_km = sum(t.distance_km for t in bus_obj.trips if t.trip_type=="Dead")
            util = rev_count / max(1, len(bus_obj.trips)) * 100

            st.markdown(
                '<div class="kpi-grid">' +
                kpi("Revenue Trips", str(rev_count)) +
                kpi("Total KM", f"{bus_obj.total_km:.1f}") +
                kpi("Dead KM", f"{dead_km:.1f}", "ok" if dead_km < 30 else "warn") +
                kpi("Final SOC", f"{bus_obj.soc_percent:.1f}%",
                    "ok" if bus_obj.soc_percent > 30 else "warn") +
                kpi("Utilisation", f"{util:.0f}%",
                    "ok" if util > 70 else "warn") +
                '</div>', unsafe_allow_html=True,
            )

            st.dataframe(bus_df, hide_index=True, height=380, use_container_width=True)

            soc_data = bus_df[bus_df["Departure"] != ""][["Departure","SOC (%)"]].set_index("Departure")
            if not soc_data.empty:
                st.markdown('<div class="section-title">SOC Progression</div>', unsafe_allow_html=True)
                st.line_chart(soc_data, color="#4f46e5")

        # ════════════════════════════════════════════════════════════════════
        # TAB 5: Config
        # ════════════════════════════════════════════════════════════════════
        with tab_config:
            # Auto-fleet detection banner
            if st.session_state.get("auto_fleet_mode"):
                detected_n = st.session_state.get("detected_fleet_size", "?")
                st.success(
                    f"🚌 **Auto Fleet Detection:** Your Excel had fleet_size = 0. "
                    f"The scheduler found that **{detected_n} buses** are the minimum required "
                    f"to serve all trips while satisfying P1–P6. "
                    f"You can lock this value below and regenerate."
                )
            if st.session_state.get("service_max_headway"):
                nat_hw = st.session_state["service_max_headway"]
                st.info(
                    f"🎯 **Service Maximization active:** Configured headway profile was replaced "
                    f"with a constant **{nat_hw:.0f}-min** headway — the minimum achievable with "
                    f"this fleet and route length. To change, adjust fleet size in the config."
                )

            st.caption("Edit any value and click **Apply & Regenerate** — no re-upload needed.")
            with st.form("config_edit_form"):
                col_a, col_b = st.columns(2)

                with col_a:
                    st.markdown("#### 🚍 Fleet & Battery")
                    new_fleet = st.number_input(
                        "Fleet Size (0 = auto-detect minimum)",
                        value=config.fleet_size, min_value=0, max_value=50, step=1,
                        help="Set to 0 to let the scheduler find the minimum fleet that satisfies all P1-P6 rules.")
                    new_battery = st.number_input("Battery (kWh)", value=float(config.battery_kwh), min_value=10.0, step=10.0)
                    new_cons    = st.number_input("Consumption (kWh/km)", value=float(config.consumption_rate), min_value=0.1, step=0.1, format="%.2f")
                    new_init_soc= st.number_input("Initial SOC (%)", value=float(config.initial_soc_percent), min_value=20.0, max_value=100.0, step=5.0)
                    new_min_km  = st.number_input("Min KM per Bus (0=none)", value=float(config.min_km_per_bus), min_value=0.0, step=10.0)
                    new_avg_spd = st.number_input(
                        "Average Speed (km/hr) — fallback when segment time is missing",
                        value=float(getattr(config, "avg_speed_kmph", 30.0)),
                        min_value=5.0, max_value=120.0, step=5.0,
                        help="Used by the distance engine to estimate travel time when a segment's time is not in the Excel.")

                    st.markdown("#### ⏰ Operating Hours")
                    new_op_start = st.text_input("Start (HH:MM)", value=config.operating_start.strftime("%H:%M"))
                    new_op_end   = st.text_input("End (HH:MM)",   value=config.operating_end.strftime("%H:%M"))
                    new_shift    = st.text_input("Shift Split (HH:MM)", value=config.shift_split.strftime("%H:%M"))

                with col_b:
                    st.markdown("#### 🔋 Charging")
                    new_depot_kw    = st.number_input("Depot Charger (kW)", value=float(config.depot_charger_kw), min_value=0.0, step=10.0)
                    new_depot_eff   = st.number_input("Depot Efficiency (0–1)", value=float(config.depot_charger_efficiency), min_value=0.0, max_value=1.0, step=0.05, format="%.2f")
                    new_trigger_soc = st.number_input(
                        "Charge Trigger SOC (%)",
                        value=float(config.trigger_soc_percent), min_value=20.0, max_value=100.0, step=5.0,
                        help="SOC level below which the scheduler sends a bus for a reactive charge stop.")
                    new_target_soc  = st.number_input("Charge Target SOC (%)", value=float(config.target_soc_percent), min_value=20.0, max_value=100.0, step=5.0)
                    new_midday_soc  = st.number_input(
                        "Midday Charge Trigger SOC (%) — P5",
                        value=float(getattr(config, "midday_charge_soc_percent", 65.0)),
                        min_value=20.0, max_value=100.0, step=5.0,
                        help="Bus must be below this SOC to be sent for midday charging (P5 window 12:00-15:00). Fleet > 10 waives this rule.")

                    st.markdown("#### 🔁 Break & Layover")
                    new_pref_layover = st.number_input(
                        "Min Driver Break (min) — P4",
                        value=int(config.preferred_layover_min), min_value=1, step=1,
                        help="Minimum break enforced between every pair of revenue trips (P4).")
                    new_max_layover  = st.number_input(
                        "Max Layover (min) — P4 upper bound",
                        value=int(getattr(config, "max_layover_min", 20)),
                        min_value=1, max_value=120, step=1,
                        help="Maximum allowed gap between revenue trips. Replaces the hardcoded 20-minute ceiling.")
                    new_off_peak_extra = st.number_input(
                        "Off-Peak Extra Break (min) — 11:00–15:00",
                        value=int(getattr(config, "off_peak_layover_extra_min", 0)),
                        min_value=0, max_value=60, step=1,
                        help="Added on top of Min Driver Break during off-peak hours to widen headways. Capped at Max Layover.")
                    new_min_layover  = st.number_input("Min Terminus Layover (min)", value=int(config.min_layover_min), min_value=0, step=1)
                    new_dead_buf     = st.number_input("Dead Run Buffer (min)", value=int(config.dead_run_buffer_min), min_value=0, step=1)
                    new_adj_buf      = st.number_input(
                        "Break Adjustment Buffer (min)",
                        value=int(config.max_headway_deviation_min), min_value=0, step=1,
                        help="Max minutes the safety-net pass can shift a trip to enforce the driver break.")

                # Segment distances (read-only)
                with st.expander("📍 Segment Distances & Times (read-only)"):
                    seg_rows = [{"Segment": k, "Distance (km)": d,
                                 "Time (min)": config.segment_times.get(k, 0)}
                                for k, d in sorted(config.segment_distances.items())]
                    st.dataframe(pd.DataFrame(seg_rows), hide_index=True)

                # Headway profile (editable) ─────────────────────────────────────
                st.markdown("#### 🕐 Headway Profile")
                st.caption(
                    "Edit **Headway (min)** per time band. "
                    "Time columns are read-only. "
                    "Changes take effect when you click Apply & Regenerate."
                )
                # Physics-based recommendations
                _hw_src_rec = st.session_state.get("raw_headway_df", pd.DataFrame())
                _tt_src_rec = st.session_state.get("raw_travel_time_df")
                if not _hw_src_rec.empty:
                    _rec_df = _headway_recommendations(config, _hw_src_rec, _tt_src_rec)
                    if _rec_df is not None:
                        with st.expander("💡 Physics-based headway recommendations", expanded=True):
                            st.caption(
                                "Recommended values are computed from your fleet size, travel times, "
                                "and break config. **⚠️ Below floor** means the configured headway "
                                "is impossible to achieve uniformly — the scheduler will silently "
                                "widen it to the Physics Floor. Change to Recommended to get "
                                "perfectly even spacing."
                            )
                            st.dataframe(
                                _rec_df,
                                hide_index=True,
                                use_container_width=True,
                                column_config={
                                    "Status": st.column_config.TextColumn(width="small"),
                                    "Note":   st.column_config.TextColumn(width="large"),
                                },
                            )
                _hw_src = st.session_state.get("raw_headway_df", pd.DataFrame())
                if not _hw_src.empty:
                    edited_hw_df = st.data_editor(
                        _hw_src[["time_from", "time_to", "headway_min"]].copy(),
                        column_config={
                            "time_from": st.column_config.TextColumn("From", disabled=True),
                            "time_to":   st.column_config.TextColumn("To",   disabled=True),
                            "headway_min": st.column_config.NumberColumn(
                                "Headway (min)", min_value=5, max_value=120, step=1,
                                help=(
                                    "Minimum gap (minutes) between consecutive same-direction "
                                    "departures from the same terminal in this time band. "
                                    "Applies fleet-wide."
                                ),
                            ),
                        },
                        hide_index=True,
                        use_container_width=True,
                        num_rows="fixed",
                        key="headway_editor",
                    )
                else:
                    edited_hw_df = None
                    st.caption("Run a schedule first to enable headway editing.")

                # ── Headway feasibility check ─────────────────────────────────
                # Warn immediately (before re-run) when a configured headway band
                # is physically impossible for the current fleet and cycle time.
                # Uses the edited headway if available, otherwise the stored one.
                _hw_for_check = edited_hw_df if edited_hw_df is not None else _hw_src
                _cfg_for_check = st.session_state.get("raw_config")
                _tt_for_check  = st.session_state.get("raw_travel_time_df")
                if (_hw_for_check is not None and not _hw_for_check.empty
                        and _cfg_for_check is not None and _tt_for_check is not None):
                    try:
                        _hw_warnings = check_headway_feasibility(
                            _cfg_for_check, _hw_for_check, _tt_for_check)
                        for _w in _hw_warnings:
                            st.warning(_w)
                    except Exception:
                        pass

                apply_btn = st.form_submit_button("🔄 Apply & Regenerate", type="primary")

            if apply_btn:
                def _pt(s):
                    try: h, m = s.strip().split(":"); return dtime(int(h), int(m))
                    except: return None

                overrides = {
                    "fleet_size": new_fleet, "battery_kwh": new_battery,
                    "consumption_rate": new_cons, "initial_soc_percent": new_init_soc,
                    "min_km_per_bus": new_min_km, "avg_speed_kmph": new_avg_spd,
                    "depot_charger_kw": new_depot_kw, "depot_charger_efficiency": new_depot_eff,
                    "trigger_soc_percent": new_trigger_soc, "target_soc_percent": new_target_soc,
                    "midday_charge_soc_percent": new_midday_soc,
                    "preferred_layover_min": new_pref_layover, "max_layover_min": new_max_layover,
                    "off_peak_layover_extra_min": new_off_peak_extra,
                    "min_layover_min": new_min_layover, "dead_run_buffer_min": new_dead_buf,
                    "max_headway_deviation_min": new_adj_buf,
                }
                for key, val in [("operating_start", new_op_start),
                                  ("operating_end", new_op_end), ("shift_split", new_shift)]:
                    t = _pt(val)
                    if t: overrides[key] = t

                hw_overrides = edited_hw_df.copy() if edited_hw_df is not None else None

                with st.spinner("Regenerating..."):
                    try:
                        result = rerun_from_overrides(overrides, headway_overrides=hw_overrides,
                                                         service_max=st.session_state.get("service_max_headway") is not None)
                    except Exception as e:
                        st.error(f"Error: {e}"); result = None

                if result:
                    new_cfg, new_buses, new_metrics, new_trips, new_out, new_comp = result
                    st.session_state["prev_compliance"] = st.session_state.get("compliance")
                    st.session_state.update({
                        "config": new_cfg, "buses": new_buses, "metrics": new_metrics,
                        "trips": new_trips, "output_bytes": new_out, "compliance": new_comp,
                    })
                    st.success("✅ Regenerated! Switch tabs to see updated results.")
                    st.rerun()



elif app_mode == "🏙️ Citywide":
    # ═══════════════════════════════ CITYWIDE ════════════════════════════════════

    if not st.session_state.get("has_city_results") and not uploaded_files:
        st.markdown("## Welcome to Citywide eBus Scheduler")
        st.markdown("Upload config Excel files for each route in the sidebar. "
                    "The system will schedule all routes and rebalance fleet across them.")
        c1, c2, c3, c4 = st.columns(4)
        icons = ["📁", "⚡", "🔄", "📊"]
        steps = [
            ("Upload", "Drop config Excel files — one per route. Same format as single-route scheduler."),
            ("Schedule", "Each route scheduled independently using its headway & travel time profiles."),
            ("Rebalance", "Surplus buses from over-allocated routes transferred to under-served routes."),
            ("Review", "Citywide KPIs, per-route drill-down, fleet transfer summary."),
        ]
        for col, (icon, (title, desc)) in zip([c1, c2, c3, c4], zip(icons, steps)):
            with col:
                st.markdown(f"### {icon} {title}")
                st.caption(desc)

    if city_run_btn and uploaded_files:
        with st.spinner("Loading configs and generating citywide schedule..."):
            try:
                city_config, load_warnings = load_city_config_from_files(uploaded_files)
                if total_fleet_override > 0:
                    city_config.total_fleet = total_fleet_override
                city_result = schedule_city(city_config, mode=city_mode)
            except ConfigError as e:
                st.error(f"Config error: {e}"); st.stop()
            except Exception as e:
                st.error(f"Error: {e}"); import traceback; st.code(traceback.format_exc()); st.stop()
        st.session_state["city_result"] = city_result
        st.session_state["city_config"] = city_config
        st.session_state["load_warnings"] = load_warnings
        st.session_state["has_city_results"] = True
        st.session_state["city_mode_used"] = city_mode

        # ── Store scenario for comparison ─────────────────────────────────────
        _scen = st.session_state.get("scenario_results", {})
        _scen[city_mode] = city_result     # overwrites previous run of same mode
        st.session_state["scenario_results"] = _scen

    if not st.session_state.get("has_city_results"):
        if uploaded_files:
            st.info("Click **▶ Generate Citywide Schedule** in the sidebar to run.")
        st.stop()

    cs: CitySchedule = st.session_state["city_result"]
    city_cfg: CityConfig = st.session_state["city_config"]
    load_warnings = st.session_state.get("load_warnings", [])

    if load_warnings:
        with st.expander(f"⚠ {len(load_warnings)} loading warning(s)", expanded=False):
            for w in load_warnings:
                st.warning(w)

    _mode_map = {
        "planning":    "📋 Planning-Compliant",
        "efficiency":  "⚡ Efficiency-Maximising",
        "service_max": "🎯 Service Maximization",
    }
    _stored_city_mode = st.session_state.get("city_mode_used", "planning")
    mode_label = _mode_map.get(_stored_city_mode, "📋 Planning-Compliant")
    st.markdown(f"## 🏙️ Citywide Schedule — {len(cs.results)} Routes")
    st.caption(f"Mode: **{mode_label}** · Depot: **{city_cfg.depot_name}**")
    _stored_city_mode = st.session_state.get("city_mode_used", "planning")
    if _stored_city_mode == "efficiency":
        st.info("⚡ **Efficiency-Maximising**: binary-searches minimum fleet per route satisfying "
                "all P1–P6 rules. Config fleet size overridden; headway floor still enforced.", icon="⚡")
    elif _stored_city_mode == "service_max":
        st.info("🎯 **Service Maximization**: config fleet used as-is. Headway profile replaced "
                "with constant minimum-achievable headway per route for even spacing.", icon="🎯")
    else:
        st.info("📋 **Planning-Compliant**: headway profile respected. "
                "Surplus buses redistributed to deficit routes based on time-sliced PVR.", icon="📋")

    # ── Citywide KPIs ─────────────────────────────────────────────────────────
    st.markdown(f"## 🏙️ Citywide Schedule — {len(cs.results)} Routes")
    st.caption(f"Mode: **{mode_label}** · Depot: **{city_cfg.depot_name}**")

    st.markdown(
        '<div class="kpi-grid">' +
        kpi("Routes", str(len(cs.results))) +
        kpi("Total Fleet", str(cs.total_buses_used),
            sub=f"configured: {city_cfg.total_configured_fleet}") +
        kpi("Revenue Trips", str(cs.total_revenue_trips)) +
        kpi("Total KM", f"{cs.total_km:,.0f}",
            sub=f"revenue: {cs.total_revenue_km:,.0f}") +
        kpi("Dead KM", f"{cs.total_dead_km:,.0f}",
            "ok" if cs.citywide_dead_km_ratio < 0.15 else "warn",
            sub=f"{cs.citywide_dead_km_ratio:.1%} of total · {cs.total_dead_trips} trips") +
        kpi("Min SOC", f"{cs.min_soc_citywide:.1f}%",
            "ok" if cs.min_soc_citywide >= 25 else "warn" if cs.min_soc_citywide >= 20 else "bad") +
        kpi("Utilization", f"{cs.citywide_utilization_pct:.0f}%",
            "ok" if cs.citywide_utilization_pct >= 85 else "warn") +
        kpi("Avg HW Dev", f"{cs.avg_headway_deviation_min:.1f} min",
            "ok" if cs.avg_headway_deviation_min < 5 else "warn"
            if cs.avg_headway_deviation_min < 10 else "bad",
            sub="vs configured") +
        '</div>', unsafe_allow_html=True,
    )

    tab_overview, tab_rebalance, tab_hw_routes, tab_depot, tab_fleet_config, tab_compare, tab_advanced = st.tabs([
        "📊 Overview", "🔄 Fleet & Rebalancing", "📈 Headways & Routes",
        "🏭 Depot", "⚙️ Config Editor", "⚖️ Compare Modes", "🔬 Advanced ▾",
    ])
    # Bind legacy variable names so existing content blocks still work
    tab_headways      = tab_hw_routes
    tab_route_detail  = tab_hw_routes
    tab_pvr           = tab_advanced
    tab_stability     = tab_advanced

    # ── CITY TAB 1: Overview ──────────────────────────────────────────────────
    with tab_overview:
        # ── Min feasible headway table (compact, no per-route error boxes) ────
        _city_mode_used = st.session_state.get("city_mode_used", "planning")
        _mf_display     = []
        _problem_count  = 0

        for _code in sorted(cs.results):
            _ri_ov   = city_cfg.routes[_code]
            # In service_max mode the scheduled headway IS the result headway, not
            # the original input config headway — use the result's headway_df
            _hw_for_check = cs.results[_code].headway_df if _city_mode_used == "service_max" else _ri_ov.headway_df
            _es = _even_spacing_min(_ri_ov.config, _hw_for_check, _ri_ov.travel_time_df)
            if not _es:
                continue
            _peak_cfg = int(_hw_for_check["headway_min"].min()) if len(_hw_for_check) else 20
            _ok       = _peak_cfg >= _es.get("peak_even_min", 0)
            if not _ok:
                _problem_count += 1
            _mf_display.append({
                "Route":              _code,
                "Peak HW (min)":      _peak_cfg,
                "Min Feasible (min)": _es.get("peak_even_min", "—"),
                "Rec k=1.0 Peak":     _es.get("peak_k10", "—"),
                "Rec k=1.0 Off-Peak": _es.get("offpeak_k10", "—"),
                "Charging RT (min)":  _es.get("charging_rt", "—"),
                "Status":             "✅ OK" if _ok else f"⚠ Need ≥{_es.get('peak_even_min', '?')}",
            })

        if _mf_display:
            if _problem_count > 0:
                st.warning(
                    f"⚠ **{_problem_count} route(s)** have configured headways below the even-spacing minimum. "
                    f"Large charging gaps are expected on these routes. See table below.",
                    icon="⚠️"
                )
            else:
                st.success("✅ All routes: configured headways meet even-spacing minimums.")

            with st.expander(
                f"📐 Minimum Feasible Headway — {_problem_count} issue(s)" if _problem_count
                else "📐 Minimum Feasible Headway — All OK",
                expanded=(_problem_count > 0)
            ):
                st.caption(
                    "**Min Feasible** = (cycle + charging_RT) ÷ fleet + 3 buffer. "
                    "**Rec k=1.0** = minimum stable. "
                    "Peak bands are those with the lowest configured headway (tighter service intent). "
                    "Off-peak = Peak + Δ (Δ = max(2, 15% of peak))."
                )
                _mf_df = pd.DataFrame(_mf_display)
                st.dataframe(_mf_df, hide_index=True, use_container_width=True,
                             column_config={"Status": st.column_config.TextColumn(width="small")})

        st.markdown('<div class="section-title">Route Summary</div>', unsafe_allow_html=True)
        summary_df = pd.DataFrame(cs.route_summary_rows())
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

        # ── Service Metrics by Route ──────────────────────────────────────────
        st.markdown('<div class="section-title">Service Metrics by Route</div>',
                    unsafe_allow_html=True)
        _svc_rows = []
        for _scode in sorted(cs.results):
            _sr  = cs.results[_scode]
            _sm  = _sr.metrics
            _all_gaps = []
            for _dir in ("UP", "DN"):
                _deps = sorted([
                    t.actual_departure for b in _sr.buses for t in b.trips
                    if t.trip_type == "Revenue" and t.direction == _dir
                    and t.actual_departure is not None
                ])
                for _i in range(1, len(_deps)):
                    _g = (_deps[_i] - _deps[_i - 1]).total_seconds() / 60
                    if _g > 0:
                        _all_gaps.append(_g)
            _avg_gap  = round(sum(_all_gaps) / len(_all_gaps)) if _all_gaps else 0
            _max_gap2 = round(getattr(_sm, 'max_headway_gap_min', 0.0))
            _ot2      = getattr(_sm, 'pct_trips_on_time', None)
            if _max_gap2 > 30:
                _interp = "⚠ Large gaps observed"
            elif _avg_gap > 0 and _max_gap2 > _avg_gap * 1.5:
                _interp = "⚠ Irregular spacing"
            else:
                _interp = "✅ Service is regular"
            _row = {
                "Route":             _scode,
                "Average Frequency": f"{_avg_gap} min" if _avg_gap else "—",
                "Longest Wait":      f"{_max_gap2} min" if _max_gap2 else "—",
                "Interpretation":    _interp,
            }
            if _ot2 is not None:
                _row["Reliable Trips %"] = f"{_ot2:.0f}%"
            _svc_rows.append(_row)
        if _svc_rows:
            st.dataframe(pd.DataFrame(_svc_rows), hide_index=True, use_container_width=True)

        if _PLOTLY_OK and len(cs.results) > 1:
            st.markdown('<div class="section-title">Fleet Allocation: PVR vs Allocated</div>',
                        unsafe_allow_html=True)
            codes = sorted(cs.results.keys())
            fig = go.Figure()
            fig.add_trace(go.Bar(name="PVR (minimum)", x=codes,
                                 y=[cs.results[c].pvr for c in codes], marker_color="#94a3b8"))
            fig.add_trace(go.Bar(name="Config Fleet", x=codes,
                                 y=[cs.results[c].fleet_original for c in codes], marker_color="#c7d2fe"))
            fig.add_trace(go.Bar(name="Final Allocated", x=codes,
                                 y=[cs.results[c].fleet_allocated for c in codes], marker_color="#4f46e5"))
            fig.update_layout(barmode="group", height=350,
                              legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0.5, xanchor="center"),
                              margin=dict(l=40, r=20, t=20, b=40), plot_bgcolor="white")
            st.plotly_chart(fig, use_container_width=True)

        if _PLOTLY_OK and len(cs.results) > 1:
            st.markdown('<div class="section-title">Dead KM % by Route</div>', unsafe_allow_html=True)
            codes = sorted(cs.results.keys())
            dead_pcts = [cs.results[c].metrics.dead_km_ratio * 100 for c in codes]
            fig2 = go.Figure(go.Bar(x=codes, y=dead_pcts, marker_color=[
                "#16a34a" if d < 15 else "#d97706" if d < 25 else "#dc2626" for d in dead_pcts]))
            fig2.update_layout(height=280, yaxis_title="Dead KM %",
                               margin=dict(l=40, r=20, t=20, b=40), plot_bgcolor="white")
            st.plotly_chart(fig2, use_container_width=True)

    # ── CITY TAB 2: Fleet & Rebalancing ───────────────────────────────────────
    with tab_rebalance:
        _city_mode_used = st.session_state.get("city_mode_used", "planning")
        balance = compute_fleet_balance(city_cfg)
        try:
            _pvr_slices_bal = compute_pvr_slices_all(city_cfg)
        except Exception:
            _pvr_slices_bal = {}

        # In planning mode: rename to "Headway Feasibility" — fleet is fixed,
        # surplus/deficit describes headway achievability, not transferable buses.
        _is_planning = (_city_mode_used == "planning")
        _section_label = "Headway Feasibility by Route" if _is_planning else "Fleet Balance Analysis (PVR-based)"
        st.markdown(f'<div class="section-title">{_section_label}</div>', unsafe_allow_html=True)

        if _is_planning:
            st.caption(
                "Fleet is fixed in Planning mode. Surplus/Deficit shows whether the "
                "configured headway is achievable with the current fleet. "
                "**Indicative only — fleet is not changed in planning mode.** "
                "A deficit means gaps are structurally unavoidable — either increase headway "
                "or add buses. See **Minimum Feasible Headway** table in Overview for specifics."
            )
        elif _city_mode_used == "service_max":
            st.caption(
                "Service Maximization uses the configured fleet as-is with a computed constant headway. "
                "Surplus/Deficit shown for reference only — no transfers are made in this mode."
            )

        bal_rows = []
        for code, b in sorted(balance.items()):
            if _city_mode_used == "efficiency" and code in _pvr_slices_bal:
                _pvr_denom = _pvr_slices_bal[code].pvr_charging
                _hr_label  = "Headroom % (vs PVR_chg)"
            else:
                _pvr_denom = b["pvr"]
                _hr_label  = "Headroom % (vs PVR_peak)"
            _hr = (b["allocated"] - _pvr_denom) / _pvr_denom * 100 if _pvr_denom > 0 else 0.0

            if _is_planning:
                # Look up even-spacing minimum fleet for this route at configured headway
                _ri_rb = city_cfg.routes.get(code)
                _es_rb = {}
                if _ri_rb:
                    try:
                        _es_rb = _even_spacing_min(_ri_rb.config, _ri_rb.headway_df, _ri_rb.travel_time_df)
                    except Exception:
                        pass
                _peak_min = _es_rb.get("peak_even_min", "—")
                _peak_cfg = int(_ri_rb.headway_df["headway_min"].min()) if _ri_rb and len(_ri_rb.headway_df) else "—"
                _hw_status = ("✅ Achievable" if isinstance(_peak_min, int) and isinstance(_peak_cfg, int)
                              and _peak_cfg >= _peak_min else
                              (f"⚠ Need ≥{_peak_min} min" if isinstance(_peak_min, int) else "—"))
                bal_rows.append({
                    "Route":           code,
                    "Fleet":           b["allocated"],
                    "PVR (min buses)": b["pvr"],
                    "Configured Peak HW": _peak_cfg,
                    "Min Feasible HW": _peak_min,
                    "HW Status":       _hw_status,
                    _hr_label:         f"{_hr:+.0f}%",
                })
            else:
                bal_rows.append({
                    "Route":    code,
                    "PVR Peak": b["pvr"],
                    "PVR Chg":  b.get("pvr_charging", b["pvr"]),
                    "Allocated": b["allocated"],
                    "Surplus":  b["surplus"] if b["surplus"] > 0 else "",
                    "Deficit":  b["deficit"] if b["deficit"] > 0 else "",
                    _hr_label:  f"{_hr:+.0f}%",
                })

        bal_df = pd.DataFrame(bal_rows)
        st.dataframe(bal_df, use_container_width=True, hide_index=True)
        st.caption(
            f"**Headroom %** uses {'PVR_charging' if _city_mode_used == 'efficiency' else 'PVR_peak'} "
            f"as denominator. Positive = spare capacity; negative = under-resourced."
        )

        # Fleet vs PVR bar chart
        if _PLOTLY_OK and len(cs.results) > 1:
            codes = sorted(cs.results.keys())
            fig_rb = go.Figure()
            fig_rb.add_trace(go.Bar(name="PVR (minimum)", x=codes,
                y=[cs.results[c].pvr for c in codes], marker_color="#94a3b8"))
            fig_rb.add_trace(go.Bar(name="Config Fleet", x=codes,
                y=[cs.results[c].fleet_original for c in codes], marker_color="#c7d2fe"))
            fig_rb.add_trace(go.Bar(name="Final Allocated", x=codes,
                y=[cs.results[c].fleet_allocated for c in codes], marker_color="#4f46e5"))
            fig_rb.update_layout(barmode="group", height=320,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0.5, xanchor="center"),
                margin=dict(l=40, r=20, t=20, b=40), plot_bgcolor="white")
            st.plotly_chart(fig_rb, use_container_width=True)

        if cs.transfers:
            st.markdown('<div class="section-title">Transfer Log</div>', unsafe_allow_html=True)
            tx_rows = [{"From": t.from_route, "To": t.to_route,
                        "Bus": t.bus_id, "Reason": t.reason}
                       for t in cs.transfers]
            st.dataframe(pd.DataFrame(tx_rows), hide_index=True, use_container_width=True)
            st.caption(f"**{len(cs.transfers)} bus(es)** redistributed to balance service.")

    # ── CITY TAB 3: Headways & Route Detail ───────────────────────────────────
    with tab_hw_routes:
        hw_route_codes = sorted(cs.results.keys())
        selected_hw_route = st.selectbox(
            "Select Route", hw_route_codes,
            format_func=lambda c: f"{c} — {cs.results[c].config.route_name}",
            key="city_hw_route",
        )
        if selected_hw_route:
            rr_hw  = cs.results[selected_hw_route]

            # ── Headway charts ────────────────────────────────────────────────
            st.markdown('<div class="section-title">Headway Chart</div>', unsafe_allow_html=True)
            hw_data = build_headway_chart_data(rr_hw.config, rr_hw.buses)
            if hw_data.empty:
                st.info("No revenue trips found for this route.")
            else:
                col_up, col_dn = st.columns(2)
                for col, direction, color in [
                    (col_up, "UP", "#4f46e5"),
                    (col_dn, "DN", "#16a34a"),
                ]:
                    with col:
                        dir_data = hw_data[hw_data["Direction"] == direction].sort_values("Departure")
                        deps = dir_data["Departure"].tolist()
                        _fig = build_headway_fig(deps, rr_hw.headway_df, direction, color,
                                                  selected_hw_route)
                        if _fig:
                            st.plotly_chart(_fig, use_container_width=True)
                        else:
                            st.caption(f"**{direction}** — fewer than 2 departures.")

            # ── Unified headway table using pre-computed RouteResult data ─────
            _rec_profile  = getattr(rr_hw, "recommended_headway_profile",  None)
            _feas_details = getattr(rr_hw, "headway_feasibility_details",   None)
            _feas_status  = getattr(rr_hw, "headway_feasibility_status", "UNKNOWN")
            _h_phys       = getattr(rr_hw, "physics_min_headway", 0)
            _rec_peak     = getattr(rr_hw, "rec_peak_headway",    0)
            _rec_offpeak  = getattr(rr_hw, "rec_offpeak_headway", 0)

            if _rec_profile or _feas_details:
                st.markdown('<div class="section-title">Headway Profile & Recommendations</div>',
                            unsafe_allow_html=True)

                # Feasibility banner
                if _feas_status == "INFEASIBLE":
                    st.error(
                        f"❌ **Headway infeasible** — one or more bands are below the physics minimum. "
                        f"Large charging gaps (80–95 min) will appear in the schedule. "
                        f"Minimum feasible peak headway: **{_h_phys} min** · "
                        f"Recommended: **{_rec_peak} min** peak · **{_rec_offpeak} min** off-peak."
                    )
                elif _feas_status == "OK":
                    st.success(
                        f"✅ All headway bands meet the physics minimum. "
                        f"Min feasible: {_h_phys} min · Rec peak: {_rec_peak} min · "
                        f"Off-peak: {_rec_offpeak} min"
                    )

                # Unified per-band table
                _unified_rows = []
                if _feas_details and _rec_profile:
                    for _fd, _rp in zip(_feas_details, _rec_profile):
                        _unified_rows.append({
                            "Band":                    _fd["band"],
                            "Input (min)":             _fd["cfg_hw"],
                            "Physics Min (min)":       _fd["physics_min"],
                            "Rec k=1.0 (min)":         _rp["headway_min"],
                            "Peak band?":              "🔵 Peak" if _rp["is_peak"] else "○ Off-peak",
                            "Status":                  _fd["status"],
                        })
                elif _feas_details:
                    for _fd in _feas_details:
                        _unified_rows.append({
                            "Band":          _fd["band"],
                            "Input (min)":   _fd["cfg_hw"],
                            "Physics Min":   _fd["physics_min"],
                            "Rec (min)":     _fd["rec"],
                            "Status":        _fd["status"],
                        })

                if _unified_rows:
                    st.dataframe(pd.DataFrame(_unified_rows), hide_index=True,
                                 use_container_width=True,
                                 column_config={"Status": st.column_config.TextColumn(width="small"),
                                                "Peak band?": st.column_config.TextColumn(width="small")})
                    st.caption(
                        "**Physics Min** = (cycle_time_for_this_band + charging_RT) ÷ fleet + "
                        f"{3} buffer (time-of-day variation applied per band).  \n"
                        "**Rec k=1.0** = minimum stable headway preserving peak < off-peak ordering. "
                        "Use Config Editor → k slider to scale up."
                    )
                    # One-click apply button with k/alpha from Config Editor sliders
                    _apply_k     = st.session_state.get(f"k_slider_{selected_hw_route}", 1.0)
                    _apply_alpha = st.session_state.get(f"alpha_slider_{selected_hw_route}", 0.15)
                    _btn_label   = (f"⚡ Apply Recommended (k={_apply_k:.2f}, α={_apply_alpha:.2f}) "
                                    f"to {selected_hw_route}")
                    if _rec_profile and st.button(_btn_label, key=f"apply_rec_{selected_hw_route}"):
                        from src.city_scheduler import _run_single_route as _rsr
                        from src.city_models import RouteInput as _RI
                        _ri_apply  = city_cfg.routes[selected_hw_route]
                        _smode     = st.session_state.get("city_mode_used", "planning")
                        with st.spinner(f"Applying recommended headways (k={_apply_k:.2f}) "
                                        f"to {selected_hw_route}…"):
                            try:
                                _new_rr = _rsr(
                                    _ri_apply,
                                    scheduling_mode=_smode,
                                    rec_k=_apply_k,
                                    rec_alpha=_apply_alpha,
                                )
                                # Apply the recommended profile as the live headway_df
                                _new_hw = pd.DataFrame([
                                    {"time_from": b["time_from"], "time_to": b["time_to"],
                                     "headway_min": b["headway_min"]}
                                    for b in _new_rr.recommended_headway_profile
                                ])
                                # Re-run schedule with the recommended headways
                                from src.trip_generator import generate_trips as _gtr
                                from src.bus_scheduler  import schedule_buses as _sbr
                                from src.metrics        import compute_metrics as _cmr
                                _trips_r = _gtr(_ri_apply.config, _new_hw, _ri_apply.travel_time_df,
                                                 scheduling_mode=_smode)
                                _buses_r = _sbr(_ri_apply.config, _trips_r, headway_df=_new_hw,
                                                 travel_time_df=_ri_apply.travel_time_df,
                                                 scheduling_mode=_smode)
                                _met_r   = _cmr(_ri_apply.config, _buses_r,
                                                 total_revenue_trips=len([t for t in _trips_r
                                                                           if t.trip_type == "Revenue"]))
                                _new_ri_r = _RI(config=_ri_apply.config, headway_df=_new_hw,
                                                 travel_time_df=_ri_apply.travel_time_df)
                                from src.fleet_analyzer import compute_pvr_slices as _cpvr2
                                _pvr_r = _cpvr2(_new_ri_r).pvr_peak
                                # Rebuild RouteResult with updated fields
                                from src.city_models import RouteResult as _RR2
                                _applied_rr = _RR2(
                                    route_code=selected_hw_route,
                                    config=_ri_apply.config,
                                    headway_df=_new_hw,
                                    travel_time_df=_ri_apply.travel_time_df,
                                    buses=_buses_r, metrics=_met_r, pvr=_pvr_r,
                                    fleet_allocated=cs.results[selected_hw_route].fleet_allocated,
                                    fleet_original=cs.results[selected_hw_route].fleet_original,
                                    physics_min_headway=_new_rr.physics_min_headway,
                                    rec_peak_headway=_new_rr.rec_peak_headway,
                                    rec_offpeak_headway=_new_rr.rec_offpeak_headway,
                                    recommended_headway_profile=_new_rr.recommended_headway_profile,
                                    headway_feasibility_status=_new_rr.headway_feasibility_status,
                                    headway_feasibility_details=_new_rr.headway_feasibility_details,
                                    headway_source=(f"recommended" if _apply_k == 1.0
                                                    else f"scaled:k{_apply_k:.2f}"),
                                    headway_k=_apply_k,
                                    headway_alpha=_apply_alpha,
                                )
                                cs.results[selected_hw_route] = _applied_rr
                                city_cfg.routes[selected_hw_route] = _new_ri_r
                                st.session_state["city_result"] = cs
                                st.session_state["city_config"] = city_cfg
                                st.success(
                                    f"✅ {selected_hw_route} rescheduled with "
                                    f"k={_apply_k:.2f} headways · "
                                    f"source: {_applied_rr.headway_source}"
                                )
                                st.rerun()
                            except Exception as _ae:
                                st.error(f"Apply error: {_ae}")

            # ── Route Detail ──────────────────────────────────────────────────
            st.divider()
            r  = rr_hw
            m  = r.metrics
            st.markdown(f'<div class="section-title">Route Detail — {r.route_code}: {r.config.route_name}</div>',
                        unsafe_allow_html=True)

            shuttle_ct = sum(1 for b in r.buses for t in b.trips if t.trip_type == "Shuttle")
            dead_ct    = sum(1 for b in r.buses for t in b.trips if t.trip_type == "Dead")
            st.markdown(
                '<div class="kpi-grid">' +
                kpi("Fleet", f"{r.fleet_allocated}",
                    sub=f"config: {r.fleet_original} · PVR: {r.pvr}") +
                kpi("Rev Trips", str(m.revenue_trips_assigned),
                    sub=f"of {m.revenue_trips_total} planned") +
                kpi("Total KM", f"{m.total_km:.0f}") +
                kpi("Dead %", f"{m.dead_km_ratio:.1%}",
                    "ok" if m.dead_km_ratio < 0.15 else "warn") +
                kpi("Min SOC", f"{m.min_soc_seen:.1f}%",
                    "ok" if m.min_soc_seen >= 25 else "warn") +
                kpi("Min Feasible HW",
                    f"{r.physics_min_headway} min" if r.physics_min_headway else "—",
                    ("ok" if (r.physics_min_headway and len(r.headway_df) > 0 and
                              int(r.headway_df['headway_min'].min()) >= r.physics_min_headway)
                     else "warn"),
                    sub=(f"rec peak: {r.rec_peak_headway} · off-pk: {r.rec_offpeak_headway}"
                         if r.rec_peak_headway else "")) +
                '</div>', unsafe_allow_html=True)

            # Download full schedule
            try:
                _sched_bytes = _build_full_schedule_excel(r.config, r.buses)
                st.download_button(
                    label=f"⬇ Download Full Schedule — {r.route_code}.xlsx",
                    data=_sched_bytes,
                    file_name=f"schedule_{r.route_code}_{r.config.route_name.replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as _dl_err:
                st.caption(f"Schedule download unavailable: {_dl_err}")

            if _PLOTLY_OK:
                st.markdown('<div class="section-title">Schedule Timeline</div>',
                            unsafe_allow_html=True)
                try:
                    gantt_fig = build_route_diagram(r.config, r.buses)
                    st.plotly_chart(gantt_fig, use_container_width=True)
                except Exception as e:
                    st.warning(f"Could not render route diagram: {e}")

            st.markdown('<div class="section-title">Bus Summary</div>', unsafe_allow_html=True)
            bus_rows = []
            for bus in r.buses:
                rev_trips = [t for t in bus.trips if t.trip_type == "Revenue"]
                bus_rows.append({
                    "Bus": bus.bus_id, "Revenue": len(rev_trips),
                    "Dead": sum(1 for t in bus.trips if t.trip_type == "Dead"),
                    "Charging": sum(1 for t in bus.trips if t.trip_type == "Charging"),
                    "Total KM": round(bus.total_km, 1),
                    "Final SOC": f"{bus.soc_percent:.1f}%",
                    "Last Location": bus.current_location,
                })
            st.dataframe(pd.DataFrame(bus_rows), use_container_width=True, hide_index=True)

            st.markdown('<div class="section-title">Full Trip Schedule</div>',
                        unsafe_allow_html=True)
            trip_rows = []
            soc_t = r.config.initial_soc_percent
            trip_seq_n = 0
            for bus in r.buses:
                for trip in bus.trips:
                    soc_t -= (trip.distance_km * r.config.consumption_rate / r.config.battery_kwh) * 100
                    if trip.trip_type == "Charging":
                        soc_t = min(100.0, soc_t + r.config.depot_flow_rate_kw *
                                    (trip.travel_time_min / 60) / r.config.battery_kwh * 100)
                    is_rev = trip.trip_type == "Revenue"
                    if is_rev:
                        trip_seq_n += 1
                    trip_rows.append({
                        "Trip #":  trip_seq_n if is_rev else "—",
                        "Bus":     trip.assigned_bus,
                        "Type":    trip.trip_type,
                        "Dir":     trip.direction,
                        "From":    trip.start_location,
                        "To":      trip.end_location,
                        "Depart":  trip.actual_departure.strftime("%H:%M") if trip.actual_departure else "",
                        "Arrive":  trip.actual_arrival.strftime("%H:%M") if trip.actual_arrival else "",
                        "KM":      round(trip.distance_km, 1),
                        "SOC %":   round(soc_t, 1),
                    })
            if trip_rows:
                st.dataframe(pd.DataFrame(trip_rows).sort_values("Depart"),
                             use_container_width=True, hide_index=True)

    # ── CITY TAB 4: Depot ─────────────────────────────────────────────────────
    with tab_depot:
        st.markdown('<div class="section-title">Depot Activity — All Routes</div>',
                    unsafe_allow_html=True)
        st.caption("Shows all buses at the depot throughout the day: arrivals, charging slots, departures, and idle time.")

        from datetime import datetime as _dt_dep, timedelta as _td_dep
        REF_D = _dt_dep(2025, 1, 1)

        # Collect all depot-related trips across all routes
        _depot_rows = []
        for _dc, _dr in cs.results.items():
            for bus in _dr.buses:
                for trip in bus.trips:
                    if trip.trip_type in ("Charging", "Dead") and (
                        trip.end_location == _dr.config.depot or
                        trip.start_location == _dr.config.depot
                    ):
                        _dep_t = trip.actual_departure
                        _arr_t = trip.actual_arrival
                        if _dep_t is None and _arr_t is None:
                            continue
                        _depot_rows.append({
                            "Route":    _dc,
                            "Bus":      trip.assigned_bus or bus.bus_id,
                            "Event":    trip.trip_type,
                            "From":     trip.start_location,
                            "To":       trip.end_location,
                            "Arrives Depot": (
                                _arr_t.strftime("%H:%M") if trip.end_location == _dr.config.depot and _arr_t
                                else "—"
                            ),
                            "Departs Depot": (
                                _dep_t.strftime("%H:%M") if trip.start_location == _dr.config.depot and _dep_t
                                else "—"
                            ),
                            "Duration (min)": trip.travel_time_min,
                        })

        if _depot_rows:
            _depot_df = (pd.DataFrame(_depot_rows)
                         .sort_values(["Arrives Depot", "Departs Depot"])
                         .reset_index(drop=True))
            st.dataframe(_depot_df, hide_index=True, use_container_width=True)

            # Concurrent charging chart
            if _PLOTLY_OK:
                st.markdown('<div class="section-title">Charging Slot Usage (30-min bins)</div>',
                            unsafe_allow_html=True)
                _bins = {}
                for _dc, _dr in cs.results.items():
                    for bus in _dr.buses:
                        for trip in bus.trips:
                            if trip.trip_type == "Charging" and trip.actual_departure and trip.actual_arrival:
                                t = trip.actual_departure
                                while t < trip.actual_arrival:
                                    _slot = t.replace(minute=(t.minute // 30) * 30, second=0)
                                    _key  = _slot.strftime("%H:%M")
                                    _bins[_key] = _bins.get(_key, 0) + 1
                                    t += _td_dep(minutes=30)

                if _bins:
                    _bin_df = pd.DataFrame(sorted(_bins.items()), columns=["Time", "Buses Charging"])
                    _peak_queue = int(_bin_df["Buses Charging"].max()) if not _bin_df.empty else 0
                    # Charger utilization: % of operating hours where ≥1 bus is charging
                    _total_slots    = len(_bin_df)
                    _active_slots   = int((_bin_df["Buses Charging"] > 0).sum())
                    _utilization_pct = round(_active_slots / _total_slots * 100) if _total_slots else 0
                    _col_qu, _col_ut = st.columns(2)
                    with _col_qu:
                        st.metric("Peak Charging Queue", f"{_peak_queue} buses",
                                  help="Max buses simultaneously at the depot charger (30-min bins).")
                    with _col_ut:
                        st.metric("Charger Utilization",  f"{_utilization_pct}%",
                                  help="% of operating hours where at least one bus is charging.")
                    fig_depot = go.Figure(go.Bar(
                        x=_bin_df["Time"], y=_bin_df["Buses Charging"],
                        marker_color="#f97316",
                        hovertemplate="%{x}: %{y} bus(es) charging<extra></extra>",
                    ))
                    fig_depot.update_layout(
                        height=260, xaxis_title="Time", yaxis_title="Concurrent Buses Charging",
                        margin=dict(l=40, r=20, t=10, b=40), plot_bgcolor="white",
                    )
                    st.plotly_chart(fig_depot, use_container_width=True)

            # Idle time summary
            st.markdown('<div class="section-title">Bus Idle Time at Depot</div>',
                        unsafe_allow_html=True)
            _idle_rows = []
            for _dc, _dr in cs.results.items():
                for bus in _dr.buses:
                    _total_idle = 0
                    for trip in bus.trips:
                        if trip.trip_type == "Charging":
                            _total_idle += trip.travel_time_min
                    _idle_rows.append({
                        "Route": _dc,
                        "Bus": bus.bus_id,
                        "Total Depot Dwell (min)": _total_idle,
                        "Charging Stops": sum(1 for t in bus.trips if t.trip_type == "Charging"),
                    })
            if _idle_rows:
                _idle_df = (pd.DataFrame(_idle_rows)
                            .sort_values("Total Depot Dwell (min)", ascending=False))
                st.dataframe(_idle_df, hide_index=True, use_container_width=True)
        else:
            st.info("No depot activity found. Buses may not have charging trips in this schedule.")
    # ── CITY TAB 6: Compare Modes ─────────────────────────────────────────────
    with tab_compare:
        _scenario_results = st.session_state.get("scenario_results", {})
        _mode_labels = {
            "planning":    "📋 Planning",
            "efficiency":  "⚡ Efficiency",
            "service_max": "🎯 Service Max",
        }

        if len(_scenario_results) < 2:
            st.info(
                "Run **at least 2 different scheduling modes** to unlock the comparison. "
                "Currently stored: " +
                (", ".join(_mode_labels.get(m, m) for m in _scenario_results)
                 if _scenario_results else "none") +
                ".  \n\n"
                "Switch mode in the sidebar and click **▶ Generate Citywide Schedule** again."
            )
        else:
            _avail_modes = [m for m in ("planning", "efficiency", "service_max")
                            if m in _scenario_results]

            # ── 1. Citywide Summary Table ─────────────────────────────────────
            st.markdown('<div class="section-title">Citywide Summary</div>',
                        unsafe_allow_html=True)

            def _scen_row(mode: str, cs_s: "CitySchedule") -> dict:
                _max_gap_s    = cs_s.max_headway_gap_min
                _on_time_s    = (sum(getattr(r.metrics, 'pct_trips_on_time', 0.0)
                                     for r in cs_s.results.values())
                                 / max(1, len(cs_s.results)))
                _infeas_count = sum(1 for r in cs_s.results.values()
                                    if getattr(r, 'headway_feasibility_status', 'UNKNOWN') == 'INFEASIBLE')
                return {
                    "Mode":               _mode_labels.get(mode, mode),
                    "Fleet Used":         cs_s.total_buses_used,
                    "Revenue Trips":      cs_s.total_revenue_trips,
                    "Max Waiting (min)":  round(_max_gap_s, 0),
                    "On-Time %":          f"{_on_time_s:.0f}%",
                    "Dead KM %":          f"{cs_s.citywide_dead_km_ratio:.1%}",
                    "Bus Utilization %":  f"{cs_s.citywide_utilization_pct:.0f}%",
                    "Min SOC %":          f"{cs_s.min_soc_citywide:.1f}%",
                    "HW Infeasible Routes": _infeas_count,
                }

            _sum_rows = [_scen_row(m, _scenario_results[m]) for m in _avail_modes]
            _sum_df   = pd.DataFrame(_sum_rows)

            # Highlight best value per metric
            _numeric_cols = ["Fleet Used", "Revenue Trips", "Max Waiting (min)",
                             "HW Infeasible Routes"]
            st.dataframe(_sum_df, hide_index=True, use_container_width=True,
                         column_config={
                             "Mode": st.column_config.TextColumn(width="medium"),
                             "Max Waiting (min)": st.column_config.NumberColumn(format="%.0f"),
                         })

            # Best/worst callouts
            _best_fleet  = min(_sum_rows, key=lambda r: r["Fleet Used"])["Mode"]
            _best_trips  = max(_sum_rows, key=lambda r: r["Revenue Trips"])["Mode"]
            _best_gap    = min(_sum_rows, key=lambda r: r["Max Waiting (min)"])["Mode"]
            _best_util   = max(_sum_rows, key=lambda r: float(r["Bus Utilization %"].rstrip('%')))["Mode"]
            _best_dead   = min(_sum_rows, key=lambda r: float(r["Dead KM %"].rstrip('%')))["Mode"]
            st.caption(
                f"🏆 **Smallest fleet:** {_best_fleet} · "
                f"**Most trips:** {_best_trips} · "
                f"**Shortest max wait:** {_best_gap} · "
                f"**Best utilization:** {_best_util} · "
                f"**Least dead KM:** {_best_dead}"
            )

            # ── 2. Visual Comparison ──────────────────────────────────────────
            if _PLOTLY_OK:
                st.markdown('<div class="section-title">Visual Comparison</div>',
                            unsafe_allow_html=True)
                _col_a, _col_b = st.columns(2)

                with _col_a:
                    # Bar: Fleet vs Revenue Trips
                    _fig_sc1 = go.Figure()
                    for _m in _avail_modes:
                        _cs_v = _scenario_results[_m]
                        _fig_sc1.add_trace(go.Bar(
                            name=_mode_labels.get(_m, _m),
                            x=["Fleet Used", "Revenue Trips"],
                            y=[_cs_v.total_buses_used, _cs_v.total_revenue_trips],
                        ))
                    _fig_sc1.update_layout(
                        barmode="group", height=280, title="Fleet vs Revenue Trips",
                        legend=dict(orientation="h", y=1.1, x=0.5, xanchor="center"),
                        margin=dict(l=40, r=20, t=40, b=40), plot_bgcolor="white",
                    )
                    st.plotly_chart(_fig_sc1, use_container_width=True)

                with _col_b:
                    # Bar: Max Waiting Time
                    _fig_sc2 = go.Figure()
                    _gap_vals  = [_scenario_results[m].max_headway_gap_min for m in _avail_modes]
                    _gap_cols  = ["#16a34a" if g <= 30 else "#d97706" if g <= 60 else "#dc2626"
                                  for g in _gap_vals]
                    _fig_sc2.add_trace(go.Bar(
                        x=[_mode_labels.get(m, m) for m in _avail_modes],
                        y=_gap_vals,
                        marker_color=_gap_cols,
                        text=[f"{g:.0f} min" for g in _gap_vals],
                        textposition="outside",
                    ))
                    _fig_sc2.update_layout(
                        height=280, title="Max Waiting Time (min)",
                        margin=dict(l=40, r=20, t=40, b=40), plot_bgcolor="white",
                        yaxis_title="minutes",
                    )
                    st.plotly_chart(_fig_sc2, use_container_width=True)

                # Radar / spider chart for normalised multi-metric comparison
                _col_c, _col_d = st.columns(2)
                with _col_c:
                    # Utilization + on-time + dead-KM bar cluster
                    _fig_sc3 = go.Figure()
                    for _m in _avail_modes:
                        _cs_v = _scenario_results[_m]
                        _ot_v = (sum(getattr(r.metrics, 'pct_trips_on_time', 0.0)
                                     for r in _cs_v.results.values())
                                 / max(1, len(_cs_v.results)))
                        _fig_sc3.add_trace(go.Bar(
                            name=_mode_labels.get(_m, _m),
                            x=["Utilization %", "On-Time %", "100 - Dead KM %"],
                            y=[
                                _cs_v.citywide_utilization_pct,
                                _ot_v,
                                100 - _cs_v.citywide_dead_km_ratio * 100,
                            ],
                        ))
                    _fig_sc3.update_layout(
                        barmode="group", height=280, title="Efficiency Metrics (higher = better)",
                        legend=dict(orientation="h", y=1.1, x=0.5, xanchor="center"),
                        yaxis=dict(range=[0, 110]),
                        margin=dict(l=40, r=20, t=40, b=40), plot_bgcolor="white",
                    )
                    st.plotly_chart(_fig_sc3, use_container_width=True)

                with _col_d:
                    # HW infeasible routes count
                    _fig_sc4 = go.Figure()
                    _infeas_vals = [
                        sum(1 for r in _scenario_results[m].results.values()
                            if getattr(r, 'headway_feasibility_status', 'OK') == 'INFEASIBLE')
                        for m in _avail_modes
                    ]
                    _fig_sc4.add_trace(go.Bar(
                        x=[_mode_labels.get(m, m) for m in _avail_modes],
                        y=_infeas_vals,
                        marker_color=["#dc2626" if v > 0 else "#16a34a" for v in _infeas_vals],
                        text=[f"{v} route(s)" for v in _infeas_vals],
                        textposition="outside",
                    ))
                    _fig_sc4.update_layout(
                        height=280, title="Infeasible Headway Routes",
                        margin=dict(l=40, r=20, t=40, b=40), plot_bgcolor="white",
                        yaxis_title="routes",
                    )
                    st.plotly_chart(_fig_sc4, use_container_width=True)

            # ── 3. Per-Route Breakdown ────────────────────────────────────────
            st.markdown('<div class="section-title">Per-Route Breakdown</div>',
                        unsafe_allow_html=True)

            _all_routes = sorted(set(
                code for cs_s in _scenario_results.values()
                for code in cs_s.results.keys()
            ))

            _pr_rows = []
            for _rc in _all_routes:
                _row = {"Route": _rc}
                for _m in _avail_modes:
                    _cs_s = _scenario_results[_m]
                    if _rc not in _cs_s.results:
                        _row[f"{_mode_labels.get(_m,'')[:3]} Fleet"] = "—"
                        _row[f"{_mode_labels.get(_m,'')[:3]} Peak HW"] = "—"
                        _row[f"{_mode_labels.get(_m,'')[:3]} Max Gap"] = "—"
                        continue
                    _rr  = _cs_s.results[_rc]
                    _mshort = _mode_labels.get(_m, _m)[:3]
                    _row[f"{_mshort} Fleet"]   = _rr.fleet_allocated
                    _row[f"{_mshort} Peak HW"] = (
                        f"{int(_rr.headway_df['headway_min'].min())} min"
                        if len(_rr.headway_df) else "—"
                    )
                    _mgap = round(getattr(_rr.metrics, 'max_headway_gap_min', 0.0))
                    _row[f"{_mshort} Max Gap"] = f"{_mgap} min"
                _pr_rows.append(_row)

            with st.expander("📋 Per-Route Detail (all modes)", expanded=False):
                st.dataframe(pd.DataFrame(_pr_rows), hide_index=True, use_container_width=True)

            # ── 4. Recommendation ─────────────────────────────────────────────
            st.markdown('<div class="section-title">💡 Recommendation</div>',
                        unsafe_allow_html=True)
            _recs = []
            if "planning" in _scenario_results and "efficiency" in _scenario_results:
                _p = _scenario_results["planning"]
                _e = _scenario_results["efficiency"]
                _fleet_saving = _p.total_buses_used - _e.total_buses_used
                _gap_diff     = _e.max_headway_gap_min - _p.max_headway_gap_min
                if _fleet_saving > 0:
                    _recs.append(
                        f"⚡ **Efficiency mode saves {_fleet_saving} bus(es)** vs Planning "
                        f"({_e.total_buses_used} vs {_p.total_buses_used} fleet)."
                    )
                if _gap_diff > 15:
                    _recs.append(
                        f"📋 **Planning mode gives better service** — max waiting time is "
                        f"{_gap_diff:.0f} min shorter than Efficiency."
                    )
            if "service_max" in _scenario_results:
                _sm = _scenario_results["service_max"]
                _infeas_sm = sum(1 for r in _sm.results.values()
                                 if getattr(r, 'headway_feasibility_status', 'OK') != 'INFEASIBLE')
                if _infeas_sm == len(_sm.results):
                    _recs.append(
                        "🎯 **Service Max** achieved even spacing on all routes — "
                        "consider it if passenger experience is the primary goal."
                    )
            if not _recs:
                _recs.append("Run all three modes to get a full recommendation.")
            for _rec in _recs:
                st.markdown(f"- {_rec}")

            # Clear button
            if st.button("🗑 Clear scenario comparison history", key="clear_scenarios"):
                st.session_state["scenario_results"] = {}
                st.rerun()

    # ── CITY TAB 7: Advanced (PVR + Stability) — collapsed expanders ──────────
    with tab_advanced:
        with st.expander("📐 PVR Analysis", expanded=False):
            st.caption(
                "**PVR (Peak)** — minimum buses needed at the tightest headway band. "
                "Formula: `ceil(cycle_time_peak / peak_headway)`.\n"
                "**PVR (Charging)** — realistic requirement including charging downtime. "
                "Formula: `ceil(PVR_peak × (1 + charging_fraction))`.\n"
                "**Off-Peak Slack** = PVR_peak − PVR_offpeak — buses freed during midday."
            )
            try:
                slices_all = compute_pvr_slices_all(city_cfg)
                pvr_rows   = [s.as_dict() for s in slices_all.values()]
                pvr_df     = pd.DataFrame(pvr_rows)
                alloc_map  = {code: r.fleet_allocated for code, r in cs.results.items()}
                pvr_df["Allocated"] = pvr_df["Route"].map(alloc_map)
                pvr_df["vs PVR Peak"] = pvr_df.apply(
                    lambda row: (
                        f"+{row['Allocated'] - row['PVR (Peak)']} surplus"
                        if row["Allocated"] > row["PVR (Peak)"]
                        else (f"{row['Allocated'] - row['PVR (Peak)']} deficit"
                              if row["Allocated"] < row["PVR (Peak)"] else "✅ exact")
                    ), axis=1
                )
                st.dataframe(pvr_df, hide_index=True, use_container_width=True)
                if _PLOTLY_OK and len(cs.results) > 1:
                    codes     = sorted(cs.results.keys())
                    fig_pvr   = go.Figure()
                    fig_pvr.add_trace(go.Bar(name="PVR (Peak)", x=codes,
                        y=[slices_all[c].pvr_peak     for c in codes], marker_color="#dc2626"))
                    fig_pvr.add_trace(go.Bar(name="PVR (Charging)", x=codes,
                        y=[slices_all[c].pvr_charging for c in codes], marker_color="#fbbf24"))
                    fig_pvr.add_trace(go.Bar(name="Allocated", x=codes,
                        y=[alloc_map.get(c, 0)        for c in codes], marker_color="#4f46e5"))
                    fig_pvr.update_layout(barmode="group", height=300,
                        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                    xanchor="center", x=0.5),
                        margin=dict(l=40, r=20, t=20, b=40), plot_bgcolor="white")
                    st.plotly_chart(fig_pvr, use_container_width=True)
            except Exception as e:
                st.warning(f"PVR analysis unavailable: {e}")

        with st.expander("🔍 Headway Stability", expanded=False):
            st.caption(
                "**On-Time %** = departures within ±3 min of ideal uniform spacing. "
                "**Stability Score** = 1/(1+std). **Avg Drift** = cumulative schedule slippage."
            )
            _stab_rows = []
            for _scode, _sr in cs.results.items():
                _sm  = _sr.metrics
                _std = getattr(_sm, "headway_std_min", 0.0)
                _ot  = getattr(_sm, "pct_trips_on_time", 0.0)
                _stab_rows.append({
                    "Route":            _scode,
                    "On-Time %":        f"{_ot:.0f}%",
                    "HW Std (min)":     round(_std, 1),
                    "Stability Score":  round(1.0 / (1.0 + _std), 3),
                    "Max Gap (min)":    round(getattr(_sm, "max_headway_gap_min", 0.0), 0),
                    "Avg Drift (min)":  round(getattr(_sm, "avg_drift_min", 0.0), 1),
                    "Max Drift (min)":  round(getattr(_sm, "max_drift_min", 0.0), 1),
                    "Quality": ("✅ Good" if _ot >= 80 and _std <= 5
                                else ("⚠️ Fair" if _ot >= 60 or _std <= 10 else "❌ Poor")),
                })
            if _stab_rows:
                st.dataframe(pd.DataFrame(_stab_rows), hide_index=True, use_container_width=True,
                    column_config={
                        "Stability Score": st.column_config.NumberColumn(format="%.3f"),
                        "HW Std (min)":    st.column_config.NumberColumn(format="%.1f"),
                    })

        with st.expander("📊 PVR Drift (Post-Rebalance)", expanded=False):
            stability_flags = getattr(cs, "stability_flags", [])
            if not stability_flags:
                st.info("PVR drift check not applicable or no transfers needed.")
            else:
                flag_rows = [f.as_dict() for f in stability_flags]
                flag_df   = pd.DataFrame(flag_rows)
                unstable  = flag_df[flag_df["Status"] == "⚠️ Drifted"]
                if not unstable.empty:
                    st.warning(f"⚠️ {len(unstable)} route(s) show PVR drift > 0.5.")
                else:
                    st.success("✅ All routes stable — no PVR drift detected.")
                st.dataframe(flag_df, hide_index=True, use_container_width=True)


    # ── CITY TAB 7: Fleet Config Editor ───────────────────────────────────────
    with tab_fleet_config:

        # ── Section A: Fleet size overview across all routes ──────────────────
        st.markdown('<div class="section-title">Fleet Size — All Routes</div>',
                    unsafe_allow_html=True)
        st.caption("Edit fleet sizes below and click **Re-run with Updated Fleet** to regenerate "
                   "the full citywide schedule with new allocations.")
        pvrs = compute_pvr_all(city_cfg)
        edit_rows = []
        for code in sorted(city_cfg.routes.keys()):
            ri = city_cfg.routes[code]
            edit_rows.append({
                "Route": code, "Name": ri.config.route_name,
                "Fleet Size": ri.config.fleet_size,
                "PVR": pvrs.get(code, 0),
                "Operating": f"{ri.config.operating_start.strftime('%H:%M')}–{ri.config.operating_end.strftime('%H:%M')}",
                "Peak Headway": int(ri.headway_df["headway_min"].min()) if len(ri.headway_df) > 0 else 30,
            })
        edit_df = pd.DataFrame(edit_rows)
        edited = st.data_editor(
            edit_df,
            column_config={
                "Route": st.column_config.TextColumn(disabled=True),
                "Name": st.column_config.TextColumn(disabled=True),
                "Fleet Size": st.column_config.NumberColumn(min_value=1, max_value=30, step=1),
                "PVR": st.column_config.NumberColumn(disabled=True),
                "Operating": st.column_config.TextColumn(disabled=True),
                "Peak Headway": st.column_config.NumberColumn(disabled=True),
            },
            use_container_width=True, hide_index=True,
        )
        total_edited = edited["Fleet Size"].sum()
        st.caption(f"**Total fleet: {total_edited}** (configured: {city_cfg.total_configured_fleet})")

        if st.button("🔄 Re-run with Updated Fleet", type="primary"):
            for _, row in edited.iterrows():
                code = row["Route"]
                new_fleet = int(row["Fleet Size"])
                if code in city_cfg.routes:
                    city_cfg.routes[code].config.fleet_size = new_fleet
            city_cfg.total_fleet = int(total_edited)
            with st.spinner("Re-running citywide schedule..."):
                try:
                    city_result = schedule_city(city_cfg, mode=st.session_state.get("city_mode_used","planning"))
                    st.session_state["city_result"] = city_result
                    st.session_state["city_config"] = city_cfg
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        st.divider()

        # ── Section B: Per-route detailed config editor ───────────────────────
        st.markdown('<div class="section-title">⚙️ Per-Route Config Editor</div>',
                    unsafe_allow_html=True)
        st.caption(
            "Select a route, edit any parameter, and click **Apply & Re-run Route**. "
            "Only the selected route is re-scheduled — all others remain unchanged."
        )

        _cfg_route_codes = sorted(city_cfg.routes.keys())
        _cfg_sel = st.selectbox(
            "Route to configure",
            _cfg_route_codes,
            format_func=lambda c: f"{c} — {city_cfg.routes[c].config.route_name}",
            key="city_cfg_route_sel",
        )

        if _cfg_sel:
            _ri_cfg  = city_cfg.routes[_cfg_sel]
            _cfg     = _ri_cfg.config
            _hw_edit = _ri_cfg.headway_df.copy()

            # ── k / Δ scaling controls (OUTSIDE form — sliders can't be inside st.form) ─
            st.markdown("#### 📐 Even-Spacing Headway Calculator")
            _es_scl   = _even_spacing_min(_cfg, _hw_edit, _ri_cfg.travel_time_df)
            _phys_min = _es_scl.get("peak_even_min", 0) if _es_scl else 0
            if _phys_min:
                st.caption(
                    f"**Physics Min:** {_phys_min} min peak · "
                    f"{_es_scl.get('offpeak_even_min','—')} min off-peak  ·  "
                    f"Charging RT ≈ {_es_scl.get('charging_rt','—')} min"
                )
            _col_k, _col_d = st.columns(2)
            with _col_k:
                _k_val = st.slider(
                    "Scaling factor k",
                    1.0, 2.0, 1.1, 0.05,
                    help="k=1.0 = minimum stable headway. k=1.1 = +10% margin. Applies to peak bands.",
                    key=f"k_slider_{_cfg_sel}"
                )
            with _col_d:
                _alpha_val = st.slider(
                    "Off-peak spread α",
                    0.05, 0.50, 0.15, 0.05,
                    help="H_offpeak = H_peak × (1 + α). α=0.15 = small gap; α=0.30 = strong gap.",
                    key=f"alpha_slider_{_cfg_sel}"
                )
            if _phys_min:
                _rec_k_peak = math.ceil(_k_val * _phys_min)
                _rec_k_offp = math.ceil(_rec_k_peak * (1.0 + _alpha_val))
                st.info(
                    f"k={_k_val:.2f}, α={_alpha_val:.2f} → "
                    f"**Peak = {_rec_k_peak} min · Off-peak = {_rec_k_offp} min** "
                    f"(diff = {_rec_k_offp - _rec_k_peak} min)"
                )
                if st.button(f"⚡ Auto-fill k={_k_val:.2f} headways into editor below",
                             key=f"autofill_{_cfg_sel}"):
                    _min_cfg_hw = int(_hw_edit["headway_min"].min()) if len(_hw_edit) else 20
                    for _idx in _hw_edit.index:
                        _is_pk = int(_hw_edit.at[_idx, "headway_min"]) <= _min_cfg_hw
                        _hw_edit.at[_idx, "headway_min"] = (
                            _rec_k_peak if _is_pk else _rec_k_offp
                        )
                    st.session_state[f"city_hw_form_edit_{_cfg_sel}"] = \
                        _hw_edit[["time_from", "time_to", "headway_min"]].copy()
                    st.rerun()

            with st.form(key=f"city_route_cfg_form_{_cfg_sel}"):
                col_a, col_b = st.columns(2)

                with col_a:
                    st.markdown("#### 🚍 Fleet & Battery")
                    _new_fleet   = st.number_input("Fleet Size", value=_cfg.fleet_size,
                                                    min_value=1, max_value=50, step=1,
                                                    help="Number of buses assigned to this route.")
                    _new_battery = st.number_input("Battery (kWh)",
                                                    value=float(_cfg.battery_kwh),
                                                    min_value=10.0, step=10.0)
                    _new_cons    = st.number_input("Consumption (kWh/km)",
                                                    value=float(_cfg.consumption_rate),
                                                    min_value=0.1, step=0.1, format="%.2f")
                    _new_init_soc = st.number_input("Initial SOC (%)",
                                                     value=float(_cfg.initial_soc_percent),
                                                     min_value=20.0, max_value=100.0, step=5.0)
                    _new_min_km  = st.number_input("Min KM per Bus (0 = none)",
                                                    value=float(getattr(_cfg, "min_km_per_bus", 0)),
                                                    min_value=0.0, step=10.0)
                    _new_avg_spd = st.number_input("Avg Speed (km/h) — fallback",
                                                    value=float(getattr(_cfg, "avg_speed_kmph", 30.0)),
                                                    min_value=5.0, max_value=120.0, step=5.0)

                    st.markdown("#### ⏰ Operating Hours")
                    _new_op_start = st.text_input("Start (HH:MM)",
                                                   value=_cfg.operating_start.strftime("%H:%M"),
                                                   key=f"op_start_{_cfg_sel}")
                    _new_op_end   = st.text_input("End (HH:MM)",
                                                   value=_cfg.operating_end.strftime("%H:%M"),
                                                   key=f"op_end_{_cfg_sel}")
                    _new_shift    = st.text_input("Shift Split (HH:MM)",
                                                   value=_cfg.shift_split.strftime("%H:%M"),
                                                   key=f"shift_{_cfg_sel}")

                with col_b:
                    st.markdown("#### 🔋 Charging")
                    _new_depot_kw   = st.number_input("Depot Charger (kW)",
                                                       value=float(_cfg.depot_charger_kw),
                                                       min_value=0.0, step=10.0,
                                                       help="Set > 0 to enable depot charging. "
                                                            "Set terminal_charger_kw > 0 in Excel to eliminate charging dead runs.")
                    _new_depot_eff  = st.number_input("Depot Charger Efficiency (0–1)",
                                                       value=float(_cfg.depot_charger_efficiency),
                                                       min_value=0.0, max_value=1.0,
                                                       step=0.05, format="%.2f")
                    _new_trig_soc   = st.number_input("Charge Trigger SOC (%)",
                                                       value=float(_cfg.trigger_soc_percent),
                                                       min_value=20.0, max_value=100.0, step=5.0,
                                                       help="SOC below which a reactive charge is triggered (P3 guard).")
                    _new_tgt_soc    = st.number_input("Charge Target SOC (%)",
                                                       value=float(_cfg.target_soc_percent),
                                                       min_value=20.0, max_value=100.0, step=5.0)
                    _new_midday_soc = st.number_input("Midday Charge Trigger SOC (%) — P5",
                                                       value=float(getattr(_cfg, "midday_charge_soc_percent", 65.0)),
                                                       min_value=20.0, max_value=100.0, step=5.0,
                                                       help="Bus must be below this SOC to qualify for P5 midday charging.")

                    st.markdown("#### 🔁 Break & Layover")
                    _new_pref_lay  = st.number_input("Min Driver Break (min) — P4",
                                                      value=int(_cfg.preferred_layover_min),
                                                      min_value=1, step=1,
                                                      help="Minimum break enforced between every pair of revenue trips.")
                    _new_max_lay   = st.number_input("Max Layover (min) — P4 upper",
                                                      value=int(getattr(_cfg, "max_layover_min", 20)),
                                                      min_value=1, max_value=120, step=1)
                    _new_opx_lay   = st.number_input("Off-Peak Extra Break (min)",
                                                      value=int(getattr(_cfg, "off_peak_layover_extra_min", 0)),
                                                      min_value=0, max_value=60, step=1,
                                                      help="Added on top of Min Driver Break during 11:00–15:00 to widen off-peak headways.")
                    _new_min_lay   = st.number_input("Min Terminus Layover (min)",
                                                      value=int(_cfg.min_layover_min),
                                                      min_value=0, step=1)
                    _new_dead_buf  = st.number_input("Dead Run Buffer (min)",
                                                      value=int(_cfg.dead_run_buffer_min),
                                                      min_value=0, step=1)
                    _new_adj_buf   = st.number_input("Break Adjustment Buffer (min)",
                                                      value=int(_cfg.max_headway_deviation_min),
                                                      min_value=0, step=1,
                                                      help="Max minutes the safety-net pass can shift a trip to enforce the driver break.")

                # Headway profile editor
                st.markdown("#### 🕐 Headway Profile")
                st.caption("Edit headway_min per time band. Time columns are read-only.")
                # Physics-based recommendations
                _rec_df_city = _headway_recommendations(_cfg, _hw_edit, _ri_cfg.travel_time_df)
                if _rec_df_city is not None:
                    with st.expander("💡 Physics-based headway recommendations", expanded=True):
                        st.caption(
                            "Recommended values are computed from fleet size, travel times, "
                            "and break config. **⚠️ Below floor** = scheduler will silently widen "
                            "to Physics Floor. Change to Recommended for perfectly even spacing."
                        )
                        st.dataframe(
                            _rec_df_city,
                            hide_index=True,
                            use_container_width=True,
                            column_config={
                                "Status": st.column_config.TextColumn(width="small"),
                                "Note":   st.column_config.TextColumn(width="large"),
                            },
                        )
                _edited_hw_city = st.data_editor(
                    _hw_edit[["time_from", "time_to", "headway_min"]].copy(),
                    column_config={
                        "time_from":   st.column_config.TextColumn("From", disabled=True),
                        "time_to":     st.column_config.TextColumn("To",   disabled=True),
                        "headway_min": st.column_config.NumberColumn(
                            "Headway (min)", min_value=5, max_value=120, step=1,
                            help="Target gap between consecutive same-direction departures in this band."),
                    },
                    hide_index=True, use_container_width=True, num_rows="fixed",
                    key=f"city_hw_form_edit_{_cfg_sel}",
                )

                # ── Headway feasibility check ─────────────────────────────────
                # Warn before re-run when any edited band is physically impossible
                # for the current fleet size and cycle time.
                try:
                    _city_hw_warnings = check_headway_feasibility(
                        _cfg, _edited_hw_city, _ri_cfg.travel_time_df)
                    for _cw in _city_hw_warnings:
                        st.warning(_cw)
                except Exception:
                    pass

                _apply_city_cfg = st.form_submit_button(
                    f"🔄 Apply & Re-run {_cfg_sel}", type="primary"
                )

            if _apply_city_cfg:
                def _parse_t(s):
                    try:
                        h, m = s.strip().split(":")
                        return dtime(int(h), int(m))
                    except Exception:
                        return None

                _overrides = {
                    "fleet_size":               _new_fleet,
                    "battery_kwh":              _new_battery,
                    "consumption_rate":         _new_cons,
                    "initial_soc_percent":      _new_init_soc,
                    "min_km_per_bus":           _new_min_km,
                    "avg_speed_kmph":           _new_avg_spd,
                    "depot_charger_kw":         _new_depot_kw,
                    "depot_charger_efficiency": _new_depot_eff,
                    "trigger_soc_percent":      _new_trig_soc,
                    "target_soc_percent":       _new_tgt_soc,
                    "midday_charge_soc_percent":_new_midday_soc,
                    "preferred_layover_min":    _new_pref_lay,
                    "max_layover_min":          _new_max_lay,
                    "off_peak_layover_extra_min": _new_opx_lay,
                    "min_layover_min":          _new_min_lay,
                    "dead_run_buffer_min":      _new_dead_buf,
                    "max_headway_deviation_min": _new_adj_buf,
                }
                for _key, _val in [("operating_start", _new_op_start),
                                    ("operating_end",   _new_op_end),
                                    ("shift_split",     _new_shift)]:
                    _t = _parse_t(_val)
                    if _t:
                        _overrides[_key] = _t

                _updated_cfg = _apply_config_overrides(_cfg, _overrides)
                _updated_hw  = _edited_hw_city.copy()

                with st.spinner(f"Re-scheduling {_cfg_sel}…"):
                    try:
                        from src.trip_generator  import generate_trips  as _gen
                        from src.bus_scheduler   import schedule_buses   as _sched
                        from src.metrics         import compute_metrics  as _metrics
                        from src.bus_scheduler   import check_compliance as _compliance
                        from src.distance_engine import enrich_distances as _enrich

                        _enrich(_updated_cfg)
                        _cfg_sched_mode = st.session_state.get("city_mode_used", "planning")
                        _trips    = _gen(_updated_cfg, _updated_hw, _ri_cfg.travel_time_df,
                                         scheduling_mode=_cfg_sched_mode)
                        _rev_ct   = len([t for t in _trips if t.trip_type == "Revenue"])
                        _buses    = _sched(_updated_cfg, _trips,
                                           headway_df=_updated_hw,
                                           travel_time_df=_ri_cfg.travel_time_df,
                                           scheduling_mode=_cfg_sched_mode)
                        _met      = _metrics(_updated_cfg, _buses, total_revenue_trips=_rev_ct)
                        _comp     = _compliance(_updated_cfg, _buses, headway_df=_updated_hw)

                        from src.city_models import RouteResult, RouteInput
                        _new_ri  = RouteInput(
                            config=_updated_cfg,
                            headway_df=_updated_hw,
                            travel_time_df=_ri_cfg.travel_time_df,
                        )
                        _new_rr  = RouteResult(
                            route_code=_cfg_sel,
                            config=_updated_cfg,
                            buses=_buses,
                            metrics=_met,
                            headway_df=_updated_hw,
                            travel_time_df=_ri_cfg.travel_time_df,
                            fleet_original=_updated_cfg.fleet_size,
                            fleet_allocated=_updated_cfg.fleet_size,
                            pvr=pvrs.get(_cfg_sel, 0),
                        )
                        cs.results[_cfg_sel]       = _new_rr
                        city_cfg.routes[_cfg_sel]  = _new_ri
                        st.session_state["city_result"] = cs
                        st.session_state["city_config"] = city_cfg
                        st.success(f"✅ {_cfg_sel} re-scheduled with updated config.")
                        # Store updated cfg for download
                        st.session_state[f"cfg_download_{_cfg_sel}"] = (
                            _updated_cfg, _updated_hw, _ri_cfg.travel_time_df)
                        st.rerun()
                    except Exception as _err:
                        st.error(f"Re-schedule error: {_err}")

            # ── Download updated config Excel (teal header = dashboard export) ─
            _dl_key = f"cfg_download_{_cfg_sel}"
            if st.session_state.get(_dl_key):
                _dl_cfg, _dl_hw, _dl_tt = st.session_state[_dl_key]
                try:
                    _cfg_bytes = _build_config_excel(_dl_cfg, _dl_hw, _dl_tt)
                    st.download_button(
                        label=f"⬇ Download Updated Config — {_cfg_sel}.xlsx",
                        data=_cfg_bytes,
                        file_name=f"ebus_config_dashboard_{_cfg_sel}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Teal header colour identifies this as a dashboard-exported config. "
                             "Upload directly on next run.",
                    )
                except Exception as _dle:
                    st.caption(f"Config download unavailable: {_dle}")


