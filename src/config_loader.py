"""
config_loader.py — Reads the eBus Config Input Excel and returns:
    1. RouteConfig object (all route parameters + segment distances)
    2. headway_df (pandas DataFrame: time_from, time_to, headway_min)
    3. travel_time_df (pandas DataFrame: time_from, time_to, up_min, dn_min)

Supports both the legacy Excel layout and the v5 template layout.

Layout differences handled transparently:

  Field map columns
    Legacy  : column B = parameter name, column C = value
    v5      : column A = parameter name, column B = value

  Travel-time sheet name
    Legacy  : "TravelTime_Profile"
    v5      : "Travel_Time"   (either is accepted)

  Segment table
    Legacy  : embedded in Route_Config, "Sr." header, resolved names in cols G/H
    v5      : standalone "Distances" sheet (A=from, B=to, C=dist_km, D=time_min)

  Coordinates
    Legacy  : embedded in Route_Config with "Location Label" header (cols B-E)
    v5      : standalone "Coordinates" sheet (A=name, B=lat, C=lon)

New fields added in v5 (read with safe defaults so legacy files continue working):
    avg_speed_kmph              default 30.0  — fallback when segment time is missing
    max_layover_min             default 20    — P4 upper break limit
    midday_charge_soc_percent   default 65.0  — P5 trigger SOC
    off_peak_layover_extra_min  default 0     — extra break 11:00-16:00
"""

from __future__ import annotations
__version__ = "2026-03-24-b1"  # auto-stamped

from datetime import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.models import RouteConfig


class ConfigError(Exception):
    """Raised when the input Excel is missing required data."""
    pass


# ── Helpers ──────────────────────────────────────────────────────────

def _parse_time(val) -> time:
    """Convert 'HH:MM' string or datetime.time to time object."""
    if isinstance(val, time):
        return val
    if isinstance(val, str):
        parts = val.strip().split(":")
        return time(int(parts[0]), int(parts[1]))
    raise ConfigError(f"Cannot parse time from: {val!r}")


def _parse_float(val, field_name: str) -> float:
    if val is None:
        raise ConfigError(f"Missing value for '{field_name}'")
    try:
        return float(val)
    except (ValueError, TypeError):
        raise ConfigError(f"Cannot parse float for '{field_name}': {val!r}")


def _parse_int(val, field_name: str) -> int:
    return int(_parse_float(val, field_name))


def _find_sheet(wb, *names: str):
    """Return the first worksheet whose name matches any of *names (case-insensitive)."""
    name_map = {s.lower(): s for s in wb.sheetnames}
    for candidate in names:
        found = name_map.get(candidate.lower())
        if found:
            return wb[found]
    return None


# ── Field-map builder ────────────────────────────────────────────────

def _build_field_map(ws) -> dict[str, tuple]:
    """
    Scan the Route_Config sheet and build {field_name: (row, value)}.

    Supports two column layouts:
      v5 template : col A = parameter name, col B = value
      Legacy v4   : col B = parameter name, col C = value

    Detection: search the first 60 rows for 'route_code' — whichever column
    contains it defines the label column.  This is reliable because 'route_code'
    is a required field that only appears in the parameter column, unlike section
    headers or legend notes that may incidentally contain underscores.
    """
    label_col = 2   # default: legacy B/C layout
    for row in range(1, min(ws.max_row + 1, 60)):
        if str(ws.cell(row, 1).value or "").strip() == "route_code":
            label_col = 1   # v5 A/B layout
            break
        if str(ws.cell(row, 2).value or "").strip() == "route_code":
            label_col = 2   # legacy B/C layout
            break
    value_col = label_col + 1

    field_map: dict[str, tuple] = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, label_col).value
        if label is None:
            continue
        label_clean = str(label).strip()
        field_map[label_clean] = (row, ws.cell(row, value_col).value)
    return field_map


def _get(field_map: dict, key: str):
    """Look up a field value by exact key. Raises ConfigError if missing."""
    if key not in field_map:
        raise ConfigError(f"Field '{key}' not found in Route_Config sheet")
    return field_map[key][1]


def _get_opt(field_map: dict, key: str, default=None):
    """Look up a field value, return default instead of raising if missing."""
    entry = field_map.get(key)
    if entry is None:
        return default
    val = entry[1]
    return val if val is not None else default


# ── Location parser ──────────────────────────────────────────────────

def _parse_locations(ws, wb, field_map: dict):
    """
    Read locations (depot, start, end, intermediates) and coordinates.

    Strategy (tries in order):
      1. Standalone "Coordinates" sheet (v5 layout)
      2. Embedded "Location Label" table in Route_Config (legacy layout)

    Returns:
        locations : dict  {label: name}
        coords    : dict  {name: (lat, lon)}
    """
    locations: dict[str, str] = {}
    coords:    dict[str, tuple] = {}

    # ── Strategy 1: standalone Coordinates sheet ─────────────────────────────
    coord_ws = _find_sheet(wb, "Coordinates", "Coords", "Locations")
    if coord_ws is not None:
        # Expect header row with "location_name" / "latitude" / "longitude"
        # or just read from row 2 onwards with A=name, B=lat, C=lon
        header_found = False
        for r in range(1, coord_ws.max_row + 1):
            v = coord_ws.cell(r, 1).value
            if v and str(v).lower() in ("location_name", "location", "name"):
                header_found = True
                continue
            if v is None:
                break
            name = str(v).strip()
            lat  = coord_ws.cell(r, 2).value
            lon  = coord_ws.cell(r, 3).value
            if name and lat is not None and lon is not None:
                try:
                    coords[name] = (float(lat), float(lon))
                except (ValueError, TypeError):
                    pass

    # Also read location labels from the field map (v5 stores them as plain fields)
    for key in ("depot", "start_point", "end_point",
                "intermediate_1", "intermediate_2"):
        val = _get_opt(field_map, key)
        if val:
            # Map to legacy location label style
            label_map = {
                "depot":          "Depot",
                "start_point":    "Start point",
                "end_point":      "End point",
                "intermediate_1": "Intermediate point 1",
                "intermediate_2": "Intermediate point 2",
            }
            locations[label_map[key]] = str(val).strip()

    # ── Strategy 2: legacy embedded "Location Label" table ───────────────────
    if not locations:
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 2).value == "Location Label":
                header_row = row
                break
        if header_row is None:
            # Also try col 1 (v5 layout)
            for row in range(1, ws.max_row + 1):
                if ws.cell(row, 1).value == "Location Label":
                    header_row = row
                    break

        if header_row is not None:
            r = header_row + 1
            while r <= ws.max_row:
                # Determine offset from detection above
                label = ws.cell(r, 2).value
                name  = ws.cell(r, 3).value
                if label is None or name is None:
                    break
                label = str(label).strip()
                name  = str(name).strip()
                locations[label] = name
                lat = ws.cell(r, 4).value
                lon = ws.cell(r, 5).value
                if lat is not None and lon is not None:
                    try:
                        coords[name] = (float(lat), float(lon))
                    except (ValueError, TypeError):
                        pass
                r += 1

    return locations, coords


# ── Segment parser ───────────────────────────────────────────────────

def _parse_segments(ws, wb, avg_speed_kmph: float = 30.0, field_map: dict = None):
    """
    Read segment distances and travel times.

    Strategy (tries in order):
      1. Standalone "Distances" sheet (v5 layout):
            A=from_location, B=to_location, C=distance_km, D=time_min
         time_min may be blank → estimated from distance / avg_speed_kmph.
      2. Embedded "Sr." table in Route_Config:
            Generic labels in cols C/D resolved via field_map or cols G/H.
            distance col E, time col F.

    Returns: segment_distances dict, segment_times dict
    """
    distances: dict[str, float] = {}
    times:     dict[str, int]   = {}

    # ── Strategy 1: standalone Distances sheet ────────────────────────────────
    dist_ws = _find_sheet(wb, "Distances", "Segment_Distances", "Segments")
    if dist_ws is not None:
        data_start = None
        for r in range(1, dist_ws.max_row + 1):
            v = dist_ws.cell(r, 1).value
            if v and str(v).lower() not in (
                "from_location", "from", "origin", "segment", "sr", "sr.",
                "segment distances & times", "segment distances and travel times",
            ):
                data_start = r
                break

        if data_start is not None:
            for r in range(data_start, dist_ws.max_row + 1):
                fr   = dist_ws.cell(r, 1).value
                to   = dist_ws.cell(r, 2).value
                dist = dist_ws.cell(r, 3).value
                tt   = dist_ws.cell(r, 4).value

                if fr is None or to is None or dist is None:
                    continue
                try:
                    fr   = str(fr).strip()
                    to   = str(to).strip()
                    dist = float(dist)
                except (ValueError, TypeError):
                    continue

                if dist <= 0:
                    continue

                if tt is None or str(tt).strip() == "":
                    tt = round(dist / max(avg_speed_kmph, 1) * 60)
                else:
                    try:
                        tt = int(float(tt))
                    except (ValueError, TypeError):
                        tt = round(dist / max(avg_speed_kmph, 1) * 60)

                key = RouteConfig.segment_key(fr, to)
                distances[key] = dist
                times[key]     = tt

        if distances:
            return distances, times
        # Fall through to legacy table if Distances sheet was empty

    # ── Strategy 2: legacy embedded "Sr." table in Route_Config ──────────────
    # Generic label → actual name resolver using field_map
    # e.g. "Depot" → field_map["depot"][1], "Start point" → field_map["start_point"][1]
    generic_to_field = {
        "Depot":                "depot",
        "Start point":          "start_point",
        "End point":            "end_point",
        "Intermediate point 1": "intermediate_1",
        "Intermediate point 2": "intermediate_2",
    }

    def _resolve(label: str) -> str | None:
        """Resolve a generic label to an actual location name."""
        label = str(label).strip()
        fkey  = generic_to_field.get(label)
        if fkey and field_map:
            val = _get_opt(field_map, fkey)
            if val:
                return str(val).strip()
        return label   # fall back to label as-is (handles absolute names)

    header_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == "Sr.":
            header_row = row
            break
    if header_row is None:
        raise ConfigError(
            "No segment data found. Expected either a 'Distances' sheet "
            "or an embedded 'Sr.' table in Route_Config."
        )

    r = header_row + 1
    while r <= ws.max_row:
        sr = ws.cell(r, 2).value
        if sr is None:
            break
        try:
            int(sr)
        except (ValueError, TypeError):
            break

        # Try resolved names from cols G/H first (older legacy format),
        # then fall back to resolving generic labels from cols C/D via field_map.
        res_from_raw = ws.cell(r, 7).value
        res_to_raw   = ws.cell(r, 8).value
        generic_from = ws.cell(r, 3).value
        generic_to   = ws.cell(r, 4).value

        if res_from_raw and res_to_raw:
            res_from = str(res_from_raw).strip()
            res_to   = str(res_to_raw).strip()
        elif generic_from and generic_to:
            res_from = _resolve(generic_from)
            res_to   = _resolve(generic_to)
        else:
            r += 1
            continue

        dist = ws.cell(r, 5).value
        tt   = ws.cell(r, 6).value

        if dist is None:
            r += 1
            continue

        try:
            dist = float(dist)
        except (ValueError, TypeError):
            r += 1
            continue

        if dist <= 0:
            r += 1
            continue

        if tt is None:
            tt = round(dist / max(avg_speed_kmph, 1) * 60)
        else:
            try:
                tt = int(float(tt))
            except (ValueError, TypeError):
                tt = round(dist / max(avg_speed_kmph, 1) * 60)

        if res_from and res_to:
            key = RouteConfig.segment_key(res_from, res_to)
            distances[key] = dist
            times[key]     = tt

        r += 1

    return distances, times


# ── Headway parser ───────────────────────────────────────────────────

def _parse_headway(wb) -> pd.DataFrame:
    """Read Headway_Profile sheet into a DataFrame."""
    ws = _find_sheet(wb, "Headway_Profile", "Headway")
    if ws is None:
        raise ConfigError("Missing required sheet: 'Headway_Profile'")

    # Find header row containing "Time From" in col B or col A
    header_row = None
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):
            if str(ws.cell(row, col).value or "").strip().lower() in ("time from", "time_from"):
                header_row = row
                # Detect which column the data starts in
                data_col = col
                break
        if header_row:
            break

    if header_row is None:
        raise ConfigError("Cannot find 'Time From' header in Headway_Profile")

    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        tf = ws.cell(r, data_col).value
        tt = ws.cell(r, data_col + 1).value
        hw = ws.cell(r, data_col + 2).value
        if tf is None:
            break
        try:
            rows.append({
                "time_from":   str(tf).strip(),
                "time_to":     str(tt).strip(),
                "headway_min": int(float(hw)),
            })
        except (ValueError, TypeError):
            pass
        r += 1

    if not rows:
        raise ConfigError("Headway_Profile sheet has no data rows")

    return pd.DataFrame(rows)


# ── Travel time parser ───────────────────────────────────────────────

def _parse_travel_time(wb) -> pd.DataFrame:
    """Read TravelTime_Profile or Travel_Time sheet into a DataFrame."""
    ws = _find_sheet(wb, "TravelTime_Profile", "Travel_Time",
                     "TravelTime", "Travel Time")
    if ws is None:
        raise ConfigError(
            "Missing required sheet: expected 'TravelTime_Profile' or 'Travel_Time'"
        )

    header_row = None
    data_col = 2
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):
            if str(ws.cell(row, col).value or "").strip().lower() in ("time from", "time_from"):
                header_row = row
                data_col = col
                break
        if header_row:
            break

    if header_row is None:
        raise ConfigError("Cannot find 'Time From' header in travel-time sheet")

    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        tf = ws.cell(r, data_col).value
        tt = ws.cell(r, data_col + 1).value
        up = ws.cell(r, data_col + 2).value
        dn = ws.cell(r, data_col + 3).value
        if tf is None:
            break
        try:
            rows.append({
                "time_from": str(tf).strip(),
                "time_to":   str(tt).strip(),
                "up_min":    int(float(up)),
                "dn_min":    int(float(dn)),
            })
        except (ValueError, TypeError):
            pass
        r += 1

    if not rows:
        raise ConfigError("Travel-time sheet has no data rows")

    return pd.DataFrame(rows)


# ── Main entry point ─────────────────────────────────────────────────

def load_config(excel_path: str | Path) -> tuple[RouteConfig, pd.DataFrame, pd.DataFrame]:
    """
    Read the eBus Config Input Excel and return:
        config          RouteConfig object
        headway_df      DataFrame with columns: time_from, time_to, headway_min
        travel_time_df  DataFrame with columns: time_from, time_to, up_min, dn_min

    Accepts both legacy (v4) and v5 template layouts transparently.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")

    wb = load_workbook(path, data_only=True)

    # ── Find the main config sheet (accept both name variants) ───────────────
    ws = _find_sheet(wb, "Route_Config", "Config", "route_config")
    if ws is None:
        raise ConfigError(
            f"Missing required sheet: expected 'Route_Config' or 'Config'. "
            f"Found sheets: {wb.sheetnames}"
        )

    # ── Field map ─────────────────────────────────────────────────────────────
    fm = _build_field_map(ws)

    # ── Read avg_speed early — needed by segment parser ───────────────────────
    avg_speed = float(_get_opt(fm, "avg_speed_kmph", 30.0) or 30.0)

    # ── Parse locations and coordinates ───────────────────────────────────────
    locations, coords = _parse_locations(ws, wb, fm)

    depot       = locations.get("Depot")
    start_point = locations.get("Start point")
    end_point   = locations.get("End point")

    if not all([depot, start_point, end_point]):
        raise ConfigError(
            f"Missing core locations. Found: Depot={depot}, "
            f"Start={start_point}, End={end_point}"
        )

    intermediates = []
    for key in ["Intermediate point 1", "Intermediate point 2"]:
        name = locations.get(key, "")
        if name:
            intermediates.append(name)

    # ── Parse segment distances and times ─────────────────────────────────────
    seg_distances, seg_times = _parse_segments(ws, wb,
                                               avg_speed_kmph=avg_speed,
                                               field_map=fm)

    if not seg_distances:
        raise ConfigError("No valid segment distances found")

    # ── Build RouteConfig ─────────────────────────────────────────────────────
    # Legacy field name for consumption rate has a unit suffix; v5 uses plain name
    consumption_key = ("consumption_rate (kWh/km)"
                       if "consumption_rate (kWh/km)" in fm
                       else "consumption_rate")

    fleet_raw = _get(fm, "fleet_size")
    fleet_size = int(float(fleet_raw)) if fleet_raw is not None else 0

    config = RouteConfig(
        route_code=str(_get(fm, "route_code")).strip(),
        route_name=str(_get(fm, "route_name")).strip(),
        depot=depot,
        start_point=start_point,
        end_point=end_point,
        intermediates=intermediates,
        fleet_size=fleet_size,
        battery_kwh=_parse_float(_get(fm, "battery_kwh"), "battery_kwh"),
        consumption_rate=_parse_float(_get(fm, consumption_key), "consumption_rate"),
        initial_soc_percent=_parse_float(
            _get(fm, "initial_soc_percent"), "initial_soc_percent"
        ),
        depot_charger_kw=_parse_float(
            _get(fm, "depot_charger_kw"), "depot_charger_kw"
        ),
        depot_charger_efficiency=_parse_float(
            _get(fm, "depot_charger_efficiency"), "depot_charger_efficiency"
        ),
        terminal_charger_kw=_parse_float(
            _get(fm, "terminal_charger_kw"), "terminal_charger_kw"
        ),
        terminal_charger_efficiency=_parse_float(
            _get(fm, "terminal_charger_efficiency"), "terminal_charger_efficiency"
        ),
        trigger_soc_percent=_parse_float(
            _get(fm, "trigger_soc_percent"), "trigger_soc_percent"
        ),
        target_soc_percent=_parse_float(
            _get(fm, "target_soc_percent"), "target_soc_percent"
        ),
        min_soc_percent=_parse_float(
            _get(fm, "min_soc_percent"), "min_soc_percent"
        ),
        min_charge_duration_min=_parse_int(
            _get(fm, "min_charge_duration_min"), "min_charge_duration_min"
        ),
        operating_start=_parse_time(_get(fm, "operating_start")),
        operating_end=_parse_time(_get(fm, "operating_end")),
        shift_split=_parse_time(_get(fm, "shift_split")),
        min_layover_min=_parse_int(
            _get(fm, "min_layover_min"), "min_layover_min"
        ),
        preferred_layover_min=_parse_int(
            _get(fm, "preferred_layover_min"), "preferred_layover_min"
        ),
        dead_run_buffer_min=_parse_int(
            _get(fm, "dead_run_buffer_min"), "dead_run_buffer_min"
        ),
        max_headway_deviation_min=_parse_int(
            _get(fm, "max_headway_deviation_min"), "max_headway_deviation_min"
        ),
        km_balance_tolerance_pct=_parse_float(
            _get(fm, "km_balance_tolerance_pct"), "km_balance_tolerance_pct"
        ),
        segment_distances=seg_distances,
        segment_times=seg_times,
        location_coords=coords,
        # ── New v5 fields — safe defaults when not present in Excel ──────────
        min_km_per_bus=float(
            _get_opt(fm, "min_km_per_bus", 0) or 0
        ),
        max_layover_min=int(float(
            _get_opt(fm, "max_layover_min", 20) or 20
        )),
        midday_charge_soc_percent=float(
            _get_opt(fm, "midday_charge_soc_percent", 65.0) or 65.0
        ),
        off_peak_layover_extra_min=int(float(
            _get_opt(fm, "off_peak_layover_extra_min", 0) or 0
        )),
        avg_speed_kmph=avg_speed,
    )

    # ── Parse headway and travel time profiles ─────────────────────────────────
    headway_df     = _parse_headway(wb)
    travel_time_df = _parse_travel_time(wb)

    wb.close()
    return config, headway_df, travel_time_df
